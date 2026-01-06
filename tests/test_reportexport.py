"""Unit tests for amc.reportexport module."""

import csv

from openpyxl import load_workbook

from amc.reportexport import export_report


class TestExportReport:
    """Tests for export_report function."""

    def test_export_report_csv_account(self, temp_output_dir):
        """Test exporting account report to CSV."""
        cost_matrix = {
            "2024-Jan": {"Account A": 100.50, "Account B": 200.75, "total": 301.25},
            "2024-Feb": {"Account A": 150.00, "Account B": 250.00, "total": 400.00},
        }
        account_list = ["Account A", "Account B", "total"]

        export_file = temp_output_dir / "test-account.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

        # Verify CSV content
        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)

            assert rows[0] == ["Month", "2024-Jan", "2024-Feb"]
            assert rows[1] == ["Account A", "100.5", "150.0"]
            assert rows[2] == ["Account B", "200.75", "250.0"]
            assert rows[3] == ["total", "301.25", "400.0"]

    def test_export_report_csv_bu(self, temp_output_dir, sample_config):
        """Test exporting business unit report to CSV."""
        cost_matrix = {
            "2024-Jan": {
                "production": 1000.00,
                "development": 500.00,
                "total": 1500.00,
            },
        }

        export_file = temp_output_dir / "test-bu.csv"
        export_report(
            export_file, cost_matrix, sample_config["account-groups"], "bu", "csv"
        )

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)

            assert rows[0] == ["Month", "2024-Jan"]
            # Should include all BUs + total
            assert len(rows) >= 3

    def test_export_report_csv_service(self, temp_output_dir):
        """Test exporting service report to CSV."""
        cost_matrix = {
            "2024-Jan": {"Amazon EC2": 800.00, "Amazon S3": 200.00, "total": 1000.00},
        }
        service_list = ["Amazon EC2", "Amazon S3", "total"]

        export_file = temp_output_dir / "test-service.csv"
        export_report(export_file, cost_matrix, service_list, "service", "csv")

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)

            assert rows[0] == ["Month", "2024-Jan"]
            assert rows[1] == ["Amazon EC2", "800.0"]

    def test_export_report_excel_account(self, temp_output_dir):
        """Test exporting account report to Excel."""
        cost_matrix = {
            "2024-Jan": {"Account A": 100.50, "Account B": 200.75, "total": 301.25},
            "2024-Feb": {"Account A": 150.00, "Account B": 250.00, "total": 400.00},
        }
        account_list = ["Account A", "Account B", "total"]

        export_file = temp_output_dir / "test-account.xlsx"
        export_report(export_file, cost_matrix, account_list, "account", "excel")

        assert export_file.exists()

        # Verify Excel content
        wb = load_workbook(export_file)
        ws = wb.active

        assert ws.cell(1, 1).value == "Month"
        assert ws.cell(1, 2).value == "2024-Jan"
        assert ws.cell(1, 3).value == "2024-Feb"
        assert ws.cell(2, 1).value == "Account A"
        assert ws.cell(2, 2).value == 100.5

    def test_export_report_excel_bu(self, temp_output_dir, sample_config):
        """Test exporting business unit report to Excel."""
        cost_matrix = {
            "2024-Jan": {
                "production": 1000.00,
                "development": 500.00,
                "total": 1500.00,
            },
        }

        export_file = temp_output_dir / "test-bu.xlsx"
        export_report(
            export_file, cost_matrix, sample_config["account-groups"], "bu", "excel"
        )

        assert export_file.exists()

        wb = load_workbook(export_file)
        ws = wb.active

        assert ws.cell(1, 1).value == "Month"
        assert ws.cell(1, 2).value == "2024-Jan"

    def test_export_report_excel_service(self, temp_output_dir):
        """Test exporting service report to Excel."""
        cost_matrix = {
            "2024-Jan": {"Amazon EC2": 800.00, "Amazon S3": 200.00, "total": 1000.00},
        }
        service_list = ["Amazon EC2", "Amazon S3", "total"]

        export_file = temp_output_dir / "test-service.xlsx"
        export_report(export_file, cost_matrix, service_list, "service", "excel")

        assert export_file.exists()

        wb = load_workbook(export_file)
        ws = wb.active

        assert ws.cell(1, 1).value == "Month"
        assert ws.cell(2, 1).value == "Amazon EC2"

    def test_export_report_creates_directory(self, tmp_path):
        """Test that export_report creates parent directory if it doesn't exist."""
        cost_matrix = {
            "2024-Jan": {"Account A": 100.00, "total": 100.00},
        }
        account_list = ["Account A", "total"]

        # Use a nested path that doesn't exist
        export_file = tmp_path / "nested" / "dir" / "test.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

    def test_export_report_zero_costs(self, temp_output_dir):
        """Test exporting report with zero costs."""
        cost_matrix = {
            "2024-Jan": {"Account A": 0.00, "total": 0.00},
        }
        account_list = ["Account A", "total"]

        export_file = temp_output_dir / "test-zero.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert rows[1] == ["Account A", "0.0"]

    def test_export_report_empty_month(self, temp_output_dir):
        """Test exporting report with missing data for some months."""
        cost_matrix = {
            "2024-Jan": {"Account A": 100.00, "total": 100.00},
            "2024-Feb": {"Account B": 200.00, "total": 200.00},  # Account A missing
        }
        account_list = ["Account A", "Account B", "total"]

        export_file = temp_output_dir / "test-missing.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
            # Account A should have 0 for Feb
            assert rows[1] == ["Account A", "100.0", "0"]

    def test_export_report_multiple_months(self, temp_output_dir):
        """Test exporting report with multiple months."""
        cost_matrix = {
            "2024-Jan": {"Account A": 100.00, "total": 100.00},
            "2024-Feb": {"Account A": 110.00, "total": 110.00},
            "2024-Mar": {"Account A": 120.00, "total": 120.00},
            "Apr": {"Account A": 130.00, "total": 130.00},
        }
        account_list = ["Account A", "total"]

        export_file = temp_output_dir / "test-multi-month.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
            # Header should have all months
            assert rows[0] == ["Month", "2024-Jan", "2024-Feb", "2024-Mar", "Apr"]

    def test_export_report_large_values(self, temp_output_dir):
        """Test exporting report with large cost values."""
        cost_matrix = {
            "2024-Jan": {"Account A": 999999.99, "total": 999999.99},
        }
        account_list = ["Account A", "total"]

        export_file = temp_output_dir / "test-large.csv"
        export_report(export_file, cost_matrix, account_list, "account", "csv")

        assert export_file.exists()

        with open(export_file, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert rows[1] == ["Account A", "999999.99"]
