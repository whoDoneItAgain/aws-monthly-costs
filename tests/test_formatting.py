"""Unit tests for amc.reportexport.formatting module."""

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from amc.reportexport.formatting import (
    CURRENCY_FORMAT,
    HEADER_ALIGNMENT_CENTER,
    HEADER_FILL_STANDARD,
    HEADER_FONT_STANDARD,
    PERCENTAGE_FORMAT,
    apply_currency_format,
    apply_header_style,
    apply_percentage_format,
    auto_adjust_column_widths,
    create_header_fill,
    create_header_font,
)


class TestConstants:
    """Tests for formatting constants."""

    def test_currency_format(self):
        """Test currency format string."""
        assert CURRENCY_FORMAT == '"$"#,##0.00'

    def test_percentage_format(self):
        """Test percentage format string."""
        assert PERCENTAGE_FORMAT == "0.00%"

    def test_header_font_standard(self):
        """Test standard header font."""
        assert isinstance(HEADER_FONT_STANDARD, Font)
        assert HEADER_FONT_STANDARD.bold is True
        assert HEADER_FONT_STANDARD.size == 14

    def test_header_fill_standard(self):
        """Test standard header fill."""
        assert isinstance(HEADER_FILL_STANDARD, PatternFill)
        assert HEADER_FILL_STANDARD.fill_type == "solid"

    def test_header_alignment_center(self):
        """Test center alignment."""
        assert isinstance(HEADER_ALIGNMENT_CENTER, Alignment)
        assert HEADER_ALIGNMENT_CENTER.horizontal == "center"


class TestCreateHeaderFont:
    """Tests for create_header_font function."""

    def test_default_parameters(self):
        """Test creating font with default parameters."""
        font = create_header_font()
        assert font.bold is True
        assert font.size == 14
        assert font.color.rgb == "FF000000"

    def test_custom_parameters(self):
        """Test creating font with custom parameters."""
        font = create_header_font(bold=False, size=16, color="FFFFFFFF")
        assert font.bold is False
        assert font.size == 16
        assert font.color.rgb == "FFFFFFFF"

    def test_only_size_custom(self):
        """Test creating font with only size customized."""
        font = create_header_font(size=18)
        assert font.bold is True
        assert font.size == 18


class TestCreateHeaderFill:
    """Tests for create_header_fill function."""

    def test_default_color(self):
        """Test creating fill with default color."""
        fill = create_header_fill()
        assert fill.start_color.rgb == "FFD9E1F2"
        assert fill.fill_type == "solid"

    def test_custom_color(self):
        """Test creating fill with custom color."""
        fill = create_header_fill(color="FF4472C4")
        assert fill.start_color.rgb == "FF4472C4"
        assert fill.fill_type == "solid"


class TestApplyHeaderStyle:
    """Tests for apply_header_style function."""

    def test_default_styling(self):
        """Test applying header style with defaults."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        result = apply_header_style(cell)
        
        assert result.font.bold is True
        assert result.fill.fill_type == "solid"
        assert result.alignment.horizontal == "center"

    def test_custom_styling(self):
        """Test applying header style with custom parameters."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        
        custom_font = Font(bold=False, size=16)
        custom_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        custom_alignment = Alignment(horizontal="left")
        
        result = apply_header_style(cell, font=custom_font, fill=custom_fill, alignment=custom_alignment)
        
        assert result.font.bold is False
        assert result.fill.start_color.rgb == "FFFFFFFF"
        assert result.alignment.horizontal == "left"


class TestApplyCurrencyFormat:
    """Tests for apply_currency_format function."""

    def test_apply_currency_format(self):
        """Test applying currency format to cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = 1234.56
        
        result = apply_currency_format(cell)
        assert result.number_format == CURRENCY_FORMAT


class TestApplyPercentageFormat:
    """Tests for apply_percentage_format function."""

    def test_apply_percentage_format(self):
        """Test applying percentage format to cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.value = 0.1234
        
        result = apply_percentage_format(cell)
        assert result.number_format == PERCENTAGE_FORMAT


class TestAutoAdjustColumnWidths:
    """Tests for auto_adjust_column_widths function."""

    def test_auto_adjust_with_content(self):
        """Test auto-adjusting column widths with content."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Short"
        ws["A2"] = "This is a much longer text"
        ws["B1"] = "Medium text"
        
        auto_adjust_column_widths(ws, max_width=50, min_width=12)
        
        # Column A should be wider due to longer text
        col_a_width = ws.column_dimensions["A"].width
        col_b_width = ws.column_dimensions["B"].width
        
        assert col_a_width >= 12  # At least min_width
        assert col_a_width <= 50  # At most max_width
        assert col_b_width >= 12

    def test_auto_adjust_empty_columns(self):
        """Test auto-adjusting with empty columns."""
        wb = Workbook()
        ws = wb.active
        
        auto_adjust_column_widths(ws, max_width=50, min_width=12)
        
        # Should apply min_width to empty columns
        col_a_width = ws.column_dimensions["A"].width
        assert col_a_width >= 12

    def test_auto_adjust_very_long_content(self):
        """Test auto-adjusting with very long content."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "A" * 100  # Very long text
        
        auto_adjust_column_widths(ws, max_width=50, min_width=12)
        
        # Should be capped at max_width
        col_a_width = ws.column_dimensions["A"].width
        assert col_a_width == 50

    def test_auto_adjust_with_none_values(self):
        """Test auto-adjusting with None values in cells."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Text"
        ws["A2"] = None
        ws["A3"] = "More text"
        
        auto_adjust_column_widths(ws, max_width=50, min_width=12)
        
        col_a_width = ws.column_dimensions["A"].width
        assert col_a_width >= 12

    def test_auto_adjust_with_numbers(self):
        """Test auto-adjusting with numeric values."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 12345.67
        ws["A2"] = 1000000
        
        auto_adjust_column_widths(ws, max_width=50, min_width=12)
        
        col_a_width = ws.column_dimensions["A"].width
        assert col_a_width >= 12
