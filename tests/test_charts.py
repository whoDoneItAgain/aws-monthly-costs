"""Unit tests for amc.reportexport.charts module."""

from openpyxl import Workbook
from openpyxl.chart import PieChart

from amc.reportexport.charts import (
    add_chart_to_worksheet,
    add_data_to_pie_chart,
    create_pie_chart,
)


class TestCreatePieChart:
    """Tests for create_pie_chart function."""

    def test_create_pie_chart_defaults(self):
        """Test creating pie chart with default parameters."""
        chart = create_pie_chart()

        assert isinstance(chart, PieChart)
        assert chart.title is None
        assert chart.height == 15
        assert chart.width == 20
        assert chart.style == 10
        assert chart.dataLabels.showCatName is True
        assert chart.dataLabels.showPercent is True
        assert chart.dataLabels.showLeaderLines is False
        assert chart.dataLabels.showVal is False
        assert chart.dataLabels.showSerName is False
        assert chart.legend is not None

    def test_create_pie_chart_with_title(self):
        """Test creating pie chart with custom title."""
        chart = create_pie_chart(title="Test Chart")

        # Chart title is an object, not a string - check it exists
        assert chart.title is not None

    def test_create_pie_chart_custom_size(self):
        """Test creating pie chart with custom size."""
        chart = create_pie_chart(height=20, width=25)

        assert chart.height == 20
        assert chart.width == 25

    def test_create_pie_chart_custom_style(self):
        """Test creating pie chart with custom style."""
        chart = create_pie_chart(style=5)

        assert chart.style == 5

    def test_create_pie_chart_no_legend(self):
        """Test creating pie chart without legend."""
        chart = create_pie_chart(show_legend=False)

        assert chart.legend is None

    def test_create_pie_chart_custom_labels(self):
        """Test creating pie chart with custom label settings."""
        chart = create_pie_chart(
            show_category_name=False,
            show_percentage=False,
            show_leader_lines=True,
            show_series_name=True,
        )

        assert chart.dataLabels.showCatName is False
        assert chart.dataLabels.showPercent is False
        assert chart.dataLabels.showLeaderLines is True
        assert chart.dataLabels.showSerName is True

    def test_create_pie_chart_all_custom_parameters(self):
        """Test creating pie chart with all custom parameters."""
        chart = create_pie_chart(
            title="Custom Chart",
            height=18,
            width=22,
            style=12,
            show_category_name=False,
            show_percentage=True,
            show_leader_lines=True,
            show_legend=False,
            show_series_name=False,
        )

        # Chart title is an object, not a string - check it exists
        assert chart.title is not None
        assert chart.height == 18
        assert chart.width == 22
        assert chart.style == 12
        assert chart.dataLabels.showCatName is False
        assert chart.dataLabels.showPercent is True
        assert chart.dataLabels.showLeaderLines is True
        assert chart.legend is None


class TestAddDataToPieChart:
    """Tests for add_data_to_pie_chart function."""

    def test_add_data_to_pie_chart(self):
        """Test adding data and labels to pie chart."""
        wb = Workbook()
        ws = wb.active

        # Add sample data
        ws["A2"] = "Category A"
        ws["A3"] = "Category B"
        ws["A4"] = "Category C"
        ws["B2"] = 100
        ws["B3"] = 200
        ws["B4"] = 300

        chart = create_pie_chart()
        result = add_data_to_pie_chart(
            chart, ws, data_col=2, label_col=1, start_row=2, end_row=4
        )

        assert result is chart
        assert len(chart.series) > 0

    def test_add_data_single_row(self):
        """Test adding data with single row."""
        wb = Workbook()
        ws = wb.active

        ws["A2"] = "Single Category"
        ws["B2"] = 500

        chart = create_pie_chart()
        result = add_data_to_pie_chart(
            chart, ws, data_col=2, label_col=1, start_row=2, end_row=2
        )

        assert result is chart

    def test_add_data_multiple_rows(self):
        """Test adding data with multiple rows."""
        wb = Workbook()
        ws = wb.active

        for i in range(2, 10):
            ws[f"A{i}"] = f"Category {i - 1}"
            ws[f"B{i}"] = i * 100

        chart = create_pie_chart()
        result = add_data_to_pie_chart(
            chart, ws, data_col=2, label_col=1, start_row=2, end_row=9
        )

        assert result is chart


class TestAddChartToWorksheet:
    """Tests for add_chart_to_worksheet function."""

    def test_add_chart_to_worksheet(self):
        """Test adding chart to worksheet."""
        wb = Workbook()
        ws = wb.active

        # Add sample data
        ws["A2"] = "Category A"
        ws["B2"] = 100

        chart = create_pie_chart(title="Test Chart")
        add_data_to_pie_chart(
            chart, ws, data_col=2, label_col=1, start_row=2, end_row=2
        )

        add_chart_to_worksheet(ws, chart, "D2")

        # Verify chart was added
        assert len(ws._charts) == 1
        assert ws._charts[0] == chart

    def test_add_multiple_charts(self):
        """Test adding multiple charts to worksheet."""
        wb = Workbook()
        ws = wb.active

        # Add sample data
        ws["A2"] = "Category A"
        ws["B2"] = 100

        chart1 = create_pie_chart(title="Chart 1")
        add_data_to_pie_chart(
            chart1, ws, data_col=2, label_col=1, start_row=2, end_row=2
        )
        add_chart_to_worksheet(ws, chart1, "D2")

        chart2 = create_pie_chart(title="Chart 2")
        add_data_to_pie_chart(
            chart2, ws, data_col=2, label_col=1, start_row=2, end_row=2
        )
        add_chart_to_worksheet(ws, chart2, "D20")

        # Verify both charts were added
        assert len(ws._charts) == 2

    def test_add_chart_different_positions(self):
        """Test adding chart at different positions."""
        wb = Workbook()
        ws = wb.active

        ws["A2"] = "Category A"
        ws["B2"] = 100

        chart = create_pie_chart()
        add_data_to_pie_chart(
            chart, ws, data_col=2, label_col=1, start_row=2, end_row=2
        )

        # Test different anchor positions
        for anchor in ["A1", "D5", "H10", "Z50"]:
            ws_test = wb.create_sheet(f"Test_{anchor}")
            ws_test["A2"] = "Category"
            ws_test["B2"] = 100
            chart_test = create_pie_chart()
            add_data_to_pie_chart(
                chart_test, ws_test, data_col=2, label_col=1, start_row=2, end_row=2
            )
            add_chart_to_worksheet(ws_test, chart_test, anchor)
            assert len(ws_test._charts) == 1
