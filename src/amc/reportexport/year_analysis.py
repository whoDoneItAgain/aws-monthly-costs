"""Year-level analysis functions for Excel reports.

This module provides functions for aggregating and analyzing AWS costs over year periods.
"""

import logging
from calendar import month_abbr, monthrange
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from amc.reportexport.analysis import _create_account_summary_sheet
from amc.reportexport.analysis_tables import (
    create_analysis_header_row,
    create_section_title,
)
from amc.reportexport.formatting import (
    add_conditional_formatting,
    auto_adjust_column_widths_advanced,
)

LOGGER = logging.getLogger(__name__)


def _aggregate_year_costs(cost_matrix: dict, year_months: list[str]) -> dict:
    """Aggregate monthly costs into yearly totals for each group.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names to aggregate (e.g., ['Jan', 'Feb', ...])

    Returns:
        Dictionary of aggregated yearly costs {group: total_cost}
    """
    year_totals = {}

    for month in year_months:
        if month in cost_matrix:
            for group, cost in cost_matrix[month].items():
                if group not in year_totals:
                    year_totals[group] = 0.0
                year_totals[group] += cost

    return {k: round(v, 2) for k, v in year_totals.items()}


def _calculate_year_daily_average(cost_matrix: dict, year_months: list[str]) -> dict:
    """Calculate daily average costs for a year period.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names in the year (e.g., ['2023-Jan', '2023-Feb', ...])

    Returns:
        Dictionary of daily average costs {group: daily_avg_cost}
    """
    month_to_num = {month_abbr[i]: i for i in range(1, 13)}

    # Calculate total days in the year period
    total_days = 0

    for month_key in year_months:
        # Extract month name from key (e.g., "2023-Jan" -> "Jan")
        if "-" in month_key:
            year_str, month_name = month_key.split("-", 1)
            year = int(year_str)
        else:
            # Fallback for month names without year
            month_name = month_key
            year = datetime.now().year

        month_num = month_to_num.get(month_name, 1)
        days_in_month = monthrange(year, month_num)[1]
        total_days += days_in_month

    # Get year totals
    year_totals = _aggregate_year_costs(cost_matrix, year_months)

    # Calculate daily average
    if total_days > 0:
        return {k: round(v / total_days, 2) for k, v in year_totals.items()}
    return year_totals


def _calculate_year_monthly_average(cost_matrix: dict, year_months: list[str]) -> dict:
    """Calculate monthly average costs for a year period.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names in the year (e.g., ['Jan', 'Feb', ...])

    Returns:
        Dictionary of monthly average costs {group: monthly_avg_cost}
    """
    # Get year totals
    year_totals = _aggregate_year_costs(cost_matrix, year_months)

    # Calculate monthly average (divide by number of months)
    num_months = len(year_months)
    if num_months > 0:
        return {k: round(v / num_months, 2) for k, v in year_totals.items()}
    return year_totals


def export_year_analysis_excel(
    output_file,
    bu_cost_matrix,
    bu_group_list,
    service_cost_matrix,
    service_group_list,
    account_cost_matrix,
    account_group_list,
    year1_months,
    year2_months,
    all_account_costs=None,
    account_id_to_name=None,
):
    """Export year-level analysis Excel file with formatted tables and charts.

    Creates year analysis tables showing:
    - Account group allocation summary (first sheet)
    - Yearly totals (last 2 complete 12-month periods comparison)
    - Daily average (for year periods)
    - Monthly average (for year periods)
    - Pie charts for business units, services, and accounts

    Args:
        output_file: Path to the output year analysis Excel file
        bu_cost_matrix: Dictionary containing BU cost data organized by month
        bu_group_list: Dictionary of BU groups
        service_cost_matrix: Dictionary containing service cost data organized by month
        service_group_list: List of services
        account_cost_matrix: Dictionary containing account cost data organized by month
        account_group_list: List of accounts
        year1_months: List of month names for first year period
        year2_months: List of month names for second year period
        all_account_costs: Dictionary of all account costs (optional, for account summary)
        account_id_to_name: Dictionary mapping account IDs to names (optional, for account summary)
    """
    LOGGER.info(f"Creating year analysis Excel file: {output_file}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Create Account Summary sheet first (if all_account_costs provided)
    if all_account_costs:
        ws_summary = wb.create_sheet("Account Summary", 0)
        # For year analysis, use both year periods for filtering
        comparison_months = year1_months + year2_months
        _create_account_summary_sheet(
            ws_summary,
            bu_group_list,
            all_account_costs,
            account_id_to_name,
            comparison_months,
        )

    # Aggregate data into yearly totals
    bu_year1 = _aggregate_year_costs(bu_cost_matrix, year1_months)
    bu_year2 = _aggregate_year_costs(bu_cost_matrix, year2_months)

    service_year1 = _aggregate_year_costs(service_cost_matrix, year1_months)
    service_year2 = _aggregate_year_costs(service_cost_matrix, year2_months)

    account_year1 = _aggregate_year_costs(account_cost_matrix, year1_months)
    account_year2 = _aggregate_year_costs(account_cost_matrix, year2_months)

    # Calculate daily averages
    bu_year1_daily = _calculate_year_daily_average(bu_cost_matrix, year1_months)
    bu_year2_daily = _calculate_year_daily_average(bu_cost_matrix, year2_months)

    service_year1_daily = _calculate_year_daily_average(
        service_cost_matrix, year1_months
    )
    service_year2_daily = _calculate_year_daily_average(
        service_cost_matrix, year2_months
    )

    account_year1_daily = _calculate_year_daily_average(
        account_cost_matrix, year1_months
    )
    account_year2_daily = _calculate_year_daily_average(
        account_cost_matrix, year2_months
    )

    # Calculate monthly averages
    bu_year1_monthly = _calculate_year_monthly_average(bu_cost_matrix, year1_months)
    bu_year2_monthly = _calculate_year_monthly_average(bu_cost_matrix, year2_months)

    service_year1_monthly = _calculate_year_monthly_average(
        service_cost_matrix, year1_months
    )
    service_year2_monthly = _calculate_year_monthly_average(
        service_cost_matrix, year2_months
    )

    account_year1_monthly = _calculate_year_monthly_average(
        account_cost_matrix, year1_months
    )
    account_year2_monthly = _calculate_year_monthly_average(
        account_cost_matrix, year2_months
    )

    # Create year labels
    year1_label = f"Year 1 ({year1_months[0]}-{year1_months[-1]})"
    year2_label = f"Year 2 ({year2_months[0]}-{year2_months[-1]})"

    # Create sheets for BU costs
    ws_bu = wb.create_sheet("BU Costs - Yearly")
    _create_year_comparison_sheet(
        ws_bu,
        "Business Unit Yearly Totals",
        bu_year1,
        bu_year2,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_BU_Year_Chart",
        chart_title="BU Costs Distribution",
    )

    ws_bu_daily = wb.create_sheet("BU Costs - Daily Avg")
    _create_year_comparison_sheet(
        ws_bu_daily,
        "Business Unit Daily Average",
        bu_year1_daily,
        bu_year2_daily,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_bu_monthly = wb.create_sheet("BU Costs - Monthly Avg")
    _create_year_comparison_sheet(
        ws_bu_monthly,
        "Business Unit Monthly Average",
        bu_year1_monthly,
        bu_year2_monthly,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # For services, calculate top 10 + Other (like monthly reports)
    # Get top 10 services by year 2 cost
    service_costs_sorted = [
        (svc, service_year2.get(svc, 0)) for svc in service_group_list if svc != "total"
    ]
    service_costs_sorted.sort(key=lambda x: x[1], reverse=True)
    top_10_services = [svc for svc, _ in service_costs_sorted[:10]]

    # Calculate "Other" for services (sum of services not in top 10)
    top_10_year1_total = sum(service_year1.get(svc, 0) for svc in top_10_services)
    top_10_year2_total = sum(service_year2.get(svc, 0) for svc in top_10_services)

    # Get BU total for calculating Other
    bu_year1_total = bu_year1.get("total", 0)
    bu_year2_total = bu_year2.get("total", 0)

    service_other_year1 = round(bu_year1_total - top_10_year1_total, 2)
    service_other_year2 = round(bu_year2_total - top_10_year2_total, 2)

    # Add Other to year1 and year2 data
    service_year1_with_other = {**service_year1, "Other": service_other_year1}
    service_year2_with_other = {**service_year2, "Other": service_other_year2}

    service_year1_daily_with_other = {
        **service_year1_daily,
        "Other": service_other_year1 / 365,
    }  # Approximate daily
    service_year2_daily_with_other = {
        **service_year2_daily,
        "Other": service_other_year2 / 365,
    }

    service_year1_monthly_with_other = {
        **service_year1_monthly,
        "Other": service_other_year1 / 12,
    }
    service_year2_monthly_with_other = {
        **service_year2_monthly,
        "Other": service_other_year2 / 12,
    }

    # Create sheets for Service costs (top 10 + Other)
    ws_service = wb.create_sheet("Top Services - Yearly")
    _create_year_comparison_sheet(
        ws_service,
        "Top Services Yearly Totals",
        service_year1_with_other,
        service_year2_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_Service_Year_Chart",
        chart_title="Top Services Distribution",
    )

    ws_service_daily = wb.create_sheet("Top Services - Daily Avg")
    _create_year_comparison_sheet(
        ws_service_daily,
        "Top Services Daily Average",
        service_year1_daily_with_other,
        service_year2_daily_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_service_monthly = wb.create_sheet("Top Services - Monthly Avg")
    _create_year_comparison_sheet(
        ws_service_monthly,
        "Top Services Monthly Average",
        service_year1_monthly_with_other,
        service_year2_monthly_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # For accounts, calculate top 10 + Other (like monthly reports)
    # Get top 10 accounts by year 2 cost
    account_costs_sorted = [
        (acc, account_year2.get(acc, 0)) for acc in account_group_list if acc != "total"
    ]
    account_costs_sorted.sort(key=lambda x: x[1], reverse=True)
    top_10_accounts = [acc for acc, _ in account_costs_sorted[:10]]

    # Calculate "Other" for accounts (sum of accounts not in top 10)
    top_10_acc_year1_total = sum(account_year1.get(acc, 0) for acc in top_10_accounts)
    top_10_acc_year2_total = sum(account_year2.get(acc, 0) for acc in top_10_accounts)

    account_other_year1 = round(bu_year1_total - top_10_acc_year1_total, 2)
    account_other_year2 = round(bu_year2_total - top_10_acc_year2_total, 2)

    # Add Other to year1 and year2 data
    account_year1_with_other = {**account_year1, "Other": account_other_year1}
    account_year2_with_other = {**account_year2, "Other": account_other_year2}

    account_year1_daily_with_other = {
        **account_year1_daily,
        "Other": account_other_year1 / 365,
    }  # Approximate daily
    account_year2_daily_with_other = {
        **account_year2_daily,
        "Other": account_other_year2 / 365,
    }

    account_year1_monthly_with_other = {
        **account_year1_monthly,
        "Other": account_other_year1 / 12,
    }
    account_year2_monthly_with_other = {
        **account_year2_monthly,
        "Other": account_other_year2 / 12,
    }

    # Create sheets for Account costs (top 10 + Other)
    ws_account = wb.create_sheet("Top Accounts - Yearly")
    _create_year_comparison_sheet(
        ws_account,
        "Top Accounts Yearly Totals",
        account_year1_with_other,
        account_year2_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_Account_Year_Chart",
        chart_title="Top Accounts Distribution",
    )

    ws_account_daily = wb.create_sheet("Top Accounts - Daily Avg")
    _create_year_comparison_sheet(
        ws_account_daily,
        "Top Accounts Daily Average",
        account_year1_daily_with_other,
        account_year2_daily_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_account_monthly = wb.create_sheet("Top Accounts - Monthly Avg")
    _create_year_comparison_sheet(
        ws_account_monthly,
        "Top Accounts Monthly Average",
        account_year1_monthly_with_other,
        account_year2_monthly_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # Save the workbook
    wb.save(output_file)
    LOGGER.info(f"Year analysis file saved: {output_file}")


def _create_year_comparison_sheet(
    worksheet,
    title,
    year1_data,
    year2_data,
    group_list,
    year1_label,
    year2_label,
    include_chart=True,
    workbook=None,
    chart_helper_name=None,
    chart_title=None,
):
    """Create a year comparison sheet with formatted table and chart.

    Args:
        worksheet: openpyxl worksheet object
        title: Title for the sheet
        year1_data: Dictionary of year 1 costs {group: cost}
        year2_data: Dictionary of year 2 costs {group: cost}
        group_list: List of groups to include
        year1_label: Label for year 1 column
        year2_label: Label for year 2 column
        include_chart: Whether to include pie chart (default: True, only for yearly totals)
        workbook: Workbook object for creating helper sheet (needed for pie chart with "Other" grouping)
        chart_helper_name: Short name for hidden helper sheet (max 31 chars, default: auto-generated)
        chart_title: Title for pie chart (default: uses year2_label Distribution)
    """
    # Title
    create_section_title(worksheet, title)

    # Headers - match monthly format with "Month" instead of "Group"
    row = 3
    headers = ["Month", year1_label, year2_label, "Difference", "% Difference"]

    # Add % Spend column header (only for sheets with totals or Other, not for daily/monthly averages)
    has_totals = "total" in group_list
    has_other = "Other" in group_list
    if (has_totals or has_other) and include_chart:
        headers.append("% Spend")

    create_analysis_header_row(worksheet, row, headers)

    # Sort groups by year2 (most recent) value in descending order, with 'Other' and 'total' at the end
    sorted_groups = []
    total_group = None
    other_group = None
    for group in group_list:
        if group == "total":
            total_group = group
        elif group == "Other":
            other_group = group
        else:
            sorted_groups.append(group)

    # Sort by most recent year's value (descending)
    sorted_groups.sort(key=lambda g: year2_data.get(g, 0), reverse=True)

    # Add Other before total (if present)
    if other_group:
        sorted_groups.append(other_group)

    # Add total at the very end if present
    if total_group:
        sorted_groups.append(total_group)

    # Data rows
    row += 1
    start_row = row

    # Calculate total for % Spend
    # For BU costs, use "total" if present
    # For Services/Accounts with "Other", calculate sum of all items including Other
    if has_totals:
        total_val2 = year2_data.get("total", 1)
    elif has_other:
        # Sum all groups including Other for % Spend calculation
        total_val2 = sum(year2_data.get(g, 0) for g in sorted_groups if g != "total")
    else:
        total_val2 = 1

    for group in sorted_groups:
        val1 = year1_data.get(group, 0)
        val2 = year2_data.get(group, 0)

        # Skip rows with zero values in both periods (except total)
        if group != "total" and val1 == 0 and val2 == 0:
            continue

        diff = val2 - val1

        # Calculate percentage difference
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        worksheet.cell(row, 1, group)
        worksheet.cell(row, 2, val1).number_format = '"$"#,##0.00'
        worksheet.cell(row, 3, val2).number_format = '"$"#,##0.00'
        # Use absolute value for difference column to match monthly format
        worksheet.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        worksheet.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        # Add % Spend column (skip for total row, but show for Other)
        if (has_totals or has_other) and include_chart and group != "total":
            pct_spend = val2 / total_val2 if total_val2 > 0 else 0
            worksheet.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    end_row = row - 1

    # Add conditional formatting
    add_conditional_formatting(
        worksheet,
        f"D{start_row}:D{end_row}",
        f"E{start_row}:E{end_row}",
    )

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(worksheet)

    # Add pie chart for year 2 data (most recent) - only if include_chart is True
    if include_chart and workbook:
        # Create helper sheet for pie chart data (groups items < 1% into "Other")
        # Use provided name or generate a short one to stay under 31 char limit
        if chart_helper_name:
            helper_sheet_name = chart_helper_name
        else:
            # Fallback to a shortened name if none provided
            helper_sheet_name = f"_Chart_{worksheet.title[:20]}"

        ws_helper = workbook.create_sheet(helper_sheet_name)
        ws_helper.sheet_state = "hidden"

        helper_row = 1
        helper_col = 1
        pie_chart_start_row = helper_row

        # Process groups in single pass: add >= 1% spend, accumulate < 1% for "Other"
        # Performance optimization: Combined 2 loops into 1 (50% reduction in iterations)
        # "Other" accumulation only happens for BU costs (not Services/Accounts with explicit Other)
        other_total = 0
        for group in sorted_groups:
            if group == "total":
                continue

            # If group is "Other" and it's in the year2_data, add it directly
            if group == "Other" and "Other" in year2_data:
                val2 = year2_data.get("Other", 0)
                if val2 > 0:
                    ws_helper.cell(helper_row, helper_col, "Other")
                    ws_helper.cell(helper_row, helper_col + 1, val2)
                    helper_row += 1
                continue

            val2 = year2_data.get(group, 0)
            val1 = year1_data.get(group, 0)

            # Skip if zero
            if val1 == 0 and val2 == 0:
                continue

            pct_spend = val2 / total_val2 if total_val2 > 0 else 0
            if pct_spend >= 0.01:  # >= 1%
                ws_helper.cell(helper_row, helper_col, group)
                ws_helper.cell(helper_row, helper_col + 1, val2)
                helper_row += 1
            elif "Other" not in year2_data:
                # Only accumulate "Other" if not already in the data
                # (i.e., only calculate for BU costs, not for Services/Accounts which have explicit Other)
                other_total += val2

        # Add accumulated "Other" if needed
        if "Other" not in year2_data and other_total > 0:
            ws_helper.cell(helper_row, helper_col, "Other")
            ws_helper.cell(helper_row, helper_col + 1, other_total)
            helper_row += 1

        pie_chart_end_row = helper_row - 1

        # Add pie chart using helper table data from hidden sheet
        if pie_chart_end_row >= pie_chart_start_row:
            chart = PieChart()
            # Remove title to prevent pie labels from falling behind it
            # The sheet already has a descriptive section title
            chart.title = None
            chart.style = 10
            # Match monthly report chart size
            chart.height = 15
            chart.width = 20

            # Use data from helper sheet
            labels = Reference(
                ws_helper,
                min_col=helper_col,
                min_row=pie_chart_start_row,
                max_row=pie_chart_end_row,
            )
            data = Reference(
                ws_helper,
                min_col=helper_col + 1,
                min_row=pie_chart_start_row,
                max_row=pie_chart_end_row,
            )

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)

            # Configure data labels to show category name and percentage only
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showCatName = True
            chart.dataLabels.showVal = False
            chart.dataLabels.showPercent = True
            chart.dataLabels.showSerName = False
            chart.legend = None

            # Position chart to match monthly reports
            worksheet.add_chart(chart, "H3")
