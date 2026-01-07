import csv
import logging
from calendar import month_abbr, monthrange
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from amc.reportexport.calculations import (
    calculate_difference,
    calculate_percentage_difference,
    calculate_percentage_spend,
)
from amc.reportexport.formatting import (
    CURRENCY_FORMAT,
    HEADER_ALIGNMENT_CENTER,
    HEADER_FILL_STANDARD,
    HEADER_FONT_STANDARD,
    PERCENTAGE_FORMAT,
    apply_currency_format,
    apply_header_style,
    apply_percentage_format,
    auto_adjust_column_widths,
)
from amc.reportexport.charts import (
    add_chart_to_worksheet,
    add_data_to_pie_chart,
    create_pie_chart,
)

LOGGER = logging.getLogger(__name__)


def _create_account_summary_sheet(worksheet, account_groups, all_account_costs):
    """Create account group allocation summary sheet.
    
    Args:
        worksheet: openpyxl worksheet object
        account_groups: Dictionary of business unit account groups
        all_account_costs: Dictionary of all account costs from API (first month used for account detection)
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get all account IDs that are defined in account_groups
    assigned_accounts = {}
    for bu, bu_accounts in account_groups.items():
        for account_id, account_name in bu_accounts.items():
            assigned_accounts[account_id] = (bu, account_name)
    
    # Get all account IDs from cost data (from the first month)
    first_month_key = next(iter(all_account_costs.keys()))
    all_cost_account_ids = set(all_account_costs[first_month_key].keys())
    
    # Identify unallocated accounts
    unallocated_account_ids = all_cost_account_ids - set(assigned_accounts.keys())
    
    # Define styles matching existing sheets
    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14, color="FF000000")
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center")
    warning_font = Font(bold=True, size=12, color="FF9C0006")
    warning_fill = PatternFill(
        start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"
    )
    
    # Title
    worksheet["A1"] = "ACCOUNT GROUP ALLOCATION SUMMARY"
    worksheet["A1"].font = title_font
    
    row = 3
    
    # Show accounts by BU
    for bu in sorted(account_groups.keys()):
        bu_accounts = account_groups[bu]
        
        # BU header with styling
        cell = worksheet.cell(row, 1, f"{bu.upper()}")
        cell.font = section_font
        row += 1
        
        # Column headers with styling matching existing sheets
        header_cell_1 = worksheet.cell(row, 1, "Account ID")
        header_cell_1.font = header_font
        header_cell_1.fill = header_fill
        header_cell_1.alignment = header_alignment
        
        header_cell_2 = worksheet.cell(row, 2, "Account Name")
        header_cell_2.font = header_font
        header_cell_2.fill = header_fill
        header_cell_2.alignment = header_alignment
        row += 1
        
        if bu_accounts:
            for account_id in sorted(bu_accounts.keys()):
                account_name_or_dict = bu_accounts[account_id]
                # Handle case where account_name might be a dict with metadata
                if isinstance(account_name_or_dict, dict):
                    # Drop metadata properties like 'cost-class', only use 'name' if present
                    account_name = account_name_or_dict.get('name', '(no name)')
                else:
                    account_name = account_name_or_dict
                worksheet.cell(row, 1, account_id)
                worksheet.cell(row, 2, str(account_name))
                row += 1
        else:
            cell = worksheet.cell(row, 1, "(no accounts)")
            cell.font = Font(italic=True, color="999999")
            row += 1
        
        row += 1  # Add blank line between BUs
    
    # Show unallocated accounts if any with warning styling
    unalloc_header = worksheet.cell(row, 1, f"UNALLOCATED ACCOUNTS ({len(unallocated_account_ids)})")
    if unallocated_account_ids:
        unalloc_header.font = warning_font
        unalloc_header.fill = warning_fill
    else:
        unalloc_header.font = section_font
    row += 1
    
    if unallocated_account_ids:
        # Column header with styling
        header_cell = worksheet.cell(row, 1, "Account ID")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        row += 1
        
        for account_id in sorted(unallocated_account_ids):
            worksheet.cell(row, 1, account_id)
            row += 1
    else:
        cell = worksheet.cell(row, 1, "None")
        cell.font = Font(italic=True, color="666666")
        row += 1
    
    # Auto-adjust column widths
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 40


def export_report(
    export_file, cost_matrix, group_list, group_by_type, output_format="csv"
):
    """Export cost report to CSV or Excel format.

    Args:
        export_file: Path to the output file
        cost_matrix: Dictionary containing cost data organized by month
        group_list: List or dictionary of groups (accounts, BUs, or services)
        group_by_type: Type of grouping ("account", "bu", or "service")
        output_format: Output format, either "csv" or "excel" (default: "csv")
    """
    # Directory should be created by caller, but ensure it exists
    export_file.parent.mkdir(parents=True, exist_ok=True)

    months = list(cost_matrix.keys())

    if output_format == "excel":
        _export_to_excel(export_file, cost_matrix, group_list, group_by_type, months)
    else:
        _export_to_csv(export_file, cost_matrix, group_list, group_by_type, months)


def _export_to_csv(export_file, cost_matrix, group_list, group_by_type, months):
    """Export cost report to CSV format."""
    csv_header = ["Month"] + months

    with open(export_file, "w", newline="") as ef:
        writer = csv.writer(ef)
        writer.writerow(csv_header)

        if group_by_type == "account":
            for account in group_list:
                csv_row = [account] + [
                    cost_matrix[month].get(account, 0) for month in months
                ]
                writer.writerow(csv_row)
        elif group_by_type == "bu":
            bus = list(group_list.keys()) + ["total"]
            for bu in bus:
                csv_row = [bu] + [cost_matrix[month].get(bu, 0) for month in months]
                writer.writerow(csv_row)
        elif group_by_type == "service":
            for service in group_list:
                csv_row = [service] + [
                    cost_matrix[month].get(service, 0) for month in months
                ]
                writer.writerow(csv_row)


def _export_to_excel(export_file, cost_matrix, group_list, group_by_type, months):
    """Export cost report to Excel format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AWS Costs"

    # Define styles for header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    # Write header
    header = ["Month"] + months
    for col_idx, header_value in enumerate(header, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header_value)
        cell.font = header_font
        cell.fill = header_fill

    # Write data rows
    row_idx = 2
    if group_by_type == "account":
        for account in group_list:
            worksheet.cell(row=row_idx, column=1, value=account)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(account, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
    elif group_by_type == "bu":
        bus = list(group_list.keys()) + ["total"]
        for bu in bus:
            worksheet.cell(row=row_idx, column=1, value=bu)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(bu, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
    elif group_by_type == "service":
        for service in group_list:
            worksheet.cell(row=row_idx, column=1, value=service)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(service, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    # Auto-adjust column widths - optimized with single pass
    for col_idx, column in enumerate(worksheet.columns, start=1):
        try:
            # Use generator expression with max() for efficiency
            max_length = max(
                (len(str(cell.value)) for cell in column if cell.value is not None),
                default=0,
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except (AttributeError, TypeError, ValueError):
            # Set default width if calculation fails
            worksheet.column_dimensions[column[0].column_letter].width = 12

    # Save workbook
    workbook.save(export_file)


def export_analysis_excel(
    output_file,
    bu_cost_matrix,
    bu_group_list,
    service_cost_matrix,
    service_group_list,
    account_cost_matrix,
    account_group_list,
    all_account_costs=None,
):
    """Export analysis Excel file with formatted tables and pie charts.

    Creates analysis tables showing:
    - Account group allocation summary (first sheet)
    - Monthly totals (last 2 months comparison)
    - Daily average (last 2 months comparison)
    - Pie charts for business units, services, and accounts

    Args:
        output_file: Path to the output analysis Excel file
        bu_cost_matrix: Dictionary containing BU cost data organized by month
        bu_group_list: Dictionary of BU groups
        service_cost_matrix: Dictionary containing service cost data organized by month
        service_group_list: List of services
        account_cost_matrix: Dictionary containing account cost data organized by month
        account_group_list: List of accounts
        all_account_costs: Dictionary of all account costs (optional, for account summary)
    """

    LOGGER.info(f"Creating analysis Excel file: {output_file}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create Account Summary sheet first (if all_account_costs provided)
    if all_account_costs:
        ws_summary = wb.create_sheet("Account Summary", 0)
        _create_account_summary_sheet(ws_summary, bu_group_list, all_account_costs)

    # Get last 2 months from the data - sort chronologically
    def get_sort_key(month_str):
        """Generate a sortable key for month strings.

        Handles both simple month format (e.g., 'Jan') and YYYY-Mon format (e.g., '2023-Jan').
        For simple format, assumes current/recent year for sorting.
        """
        try:
            # Try simple month format first (e.g., "Jan")
            month_date = datetime.strptime(month_str, "%b")
            # For simple format, use a recent year for sorting
            return datetime(2024, month_date.month, 1)
        except ValueError:
            try:
                # Try YYYY-Mon format (e.g., "2023-Jan")
                return datetime.strptime(month_str, "%Y-%b")
            except ValueError:
                # Fallback to string sorting
                return datetime.min

    months = sorted(list(bu_cost_matrix.keys()), key=get_sort_key)
    if len(months) < 2:
        LOGGER.warning("Need at least 2 months of data for analysis")
        return

    last_2_months = months[-2:]

    # Create analysis sheets
    ws_bu = wb.create_sheet("BU Costs")
    ws_bu_daily = wb.create_sheet("BU Daily Average")
    ws_bu_helper = wb.create_sheet("_BU_Chart_Data")
    ws_bu_helper.sheet_state = "hidden"  # Hide the helper sheet
    _create_bu_analysis_tables(
        ws_bu, ws_bu_daily, ws_bu_helper, bu_cost_matrix, bu_group_list, last_2_months
    )

    ws_service = wb.create_sheet("Top Services")
    ws_service_daily = wb.create_sheet("Top Services Daily Avg")
    _create_service_analysis_tables(
        ws_service,
        ws_service_daily,
        service_cost_matrix,
        service_group_list,
        last_2_months,
        bu_cost_matrix,
    )

    ws_account = wb.create_sheet("Top Accounts")
    ws_account_daily = wb.create_sheet("Top Accounts Daily Avg")
    _create_account_analysis_tables(
        ws_account,
        ws_account_daily,
        account_cost_matrix,
        account_group_list,
        last_2_months,
        bu_cost_matrix,
    )

    # Ensure output directory exists
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Save the workbook
    wb.save(output_file)
    LOGGER.info(f"Analysis Excel file saved: {output_file}")


def _create_bu_analysis_tables(
    ws, ws_daily, ws_helper, cost_matrix, group_list, last_2_months
):
    """Create BU analysis tables with monthly totals on one sheet, daily average on another, and pie chart with helper table in hidden sheet."""
    # Monthly Totals section
    ws["A1"] = "BU Monthly Totals"
    ws["A1"].font = Font(bold=True, size=16)

    row = 3
    # Apply header styling using utility functions
    for col, header_text in enumerate(
        [
            "Month",
            last_2_months[0],
            last_2_months[1],
            "Difference",
            "% Difference",
            "% Spend",
        ],
        start=1,
    ):
        apply_header_style(ws.cell(row, col, header_text))

    # Data rows for monthly totals
    row += 1
    data_start_row = row

    # Sort BUs by most recent month's cost in descending order
    bus = list(group_list.keys())
    bus.sort(key=lambda bu: cost_matrix[last_2_months[1]].get(bu, 0), reverse=True)

    # Cache month dictionaries for faster lookups (performance optimization)
    # These cached references eliminate repeated dictionary traversals in the loops below
    # for main table, daily average calculations, and chart data processing
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]
    month2_total = month2_costs.get("total", 1)

    # Display all BUs in the table
    for bu in bus:
        val1 = month1_costs.get(bu, 0)
        val2 = month2_costs.get(bu, 0)

        # Skip this BU if both current and previous months are zero
        if val1 == 0 and val2 == 0:
            continue

        ws.cell(row, 1, bu)

        # Use calculation utilities
        diff = calculate_difference(val1, val2)
        pct_diff = calculate_percentage_difference(val1, val2)
        pct_spend = calculate_percentage_spend(val2, month2_total)

        # Apply formatting using utility functions
        apply_currency_format(ws.cell(row, 2, val1))
        apply_currency_format(ws.cell(row, 3, val2))
        apply_currency_format(ws.cell(row, 4, diff))
        apply_percentage_format(ws.cell(row, 5, abs(pct_diff)))
        apply_percentage_format(ws.cell(row, 6, pct_spend))

        row += 1

    # Total row
    total1 = month1_costs.get("total", 0)
    total2 = month2_costs.get("total", 0)
    diff = calculate_difference(total1, total2)
    pct_diff = calculate_percentage_difference(total1, total2)

    ws.cell(row, 1, "total")
    apply_currency_format(ws.cell(row, 2, total1))
    apply_currency_format(ws.cell(row, 3, total2))
    apply_currency_format(ws.cell(row, 4, diff))
    apply_percentage_format(ws.cell(row, 5, abs(pct_diff)))
    # Column 6 (% Spend) intentionally left empty for total row - it's implied to be 100%

    data_end_row = row

    # Add conditional formatting to difference and % difference columns
    _add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Create helper table for pie chart in hidden sheet (groups BUs < 1% into "Other")
    helper_row = 1
    helper_col = 1  # Column A in helper sheet

    total_for_pct = month2_total  # Reuse cached value
    pie_chart_start_row = helper_row

    # Process BUs in single pass: add >= 1% spend, accumulate < 1% for "Other"
    # Performance optimization: Combined 2 loops into 1 (50% reduction in iterations)
    other_total = 0
    for bu in bus:
        val2 = month2_costs.get(bu, 0)
        val1 = month1_costs.get(bu, 0)

        # Skip if zero
        if val1 == 0 and val2 == 0:
            continue

        pct_spend = val2 / total_for_pct
        if pct_spend >= 0.01:  # >= 1%
            ws_helper.cell(helper_row, helper_col, bu)
            ws_helper.cell(helper_row, helper_col + 1, val2)
            helper_row += 1
        else:  # < 1%
            other_total += val2

    if other_total > 0:
        ws_helper.cell(helper_row, helper_col, "Other")
        ws_helper.cell(helper_row, helper_col + 1, other_total)
        helper_row += 1

    pie_chart_end_row = helper_row - 1

    # Add pie chart using helper table data from hidden sheet
    # Only add pie chart if there's BU data to display
    if pie_chart_end_row >= pie_chart_start_row:
        chart = create_pie_chart(show_legend=False, show_series_name=False)
        chart = add_data_to_pie_chart(
            chart,
            ws_helper,
            helper_col + 1,
            helper_col,
            pie_chart_start_row,
            pie_chart_end_row,
        )
        add_chart_to_worksheet(ws, chart, "H3")

    # Auto-adjust column widths
    auto_adjust_column_widths(ws)

    # Daily Average section - on separate sheet
    ws_daily["A1"] = "BU Daily Average"
    ws_daily["A1"].font = Font(bold=True, size=16)

    row = 3

    # Apply header styling using utility functions
    for col, header_text in enumerate(
        ["Month", last_2_months[0], last_2_months[1], "Difference", "% Difference"],
        start=1,
    ):
        apply_header_style(ws_daily.cell(row, col, header_text))

    # Calculate days in each month
    # For the last 2 months, infer the year intelligently
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        # Fallback to 30 days if parsing fails
        days1 = days2 = 30

    # Data rows for daily average
    row += 1
    daily_start_row = row
    for bu in bus:
        val1_monthly = month1_costs.get(bu, 0)
        val2_monthly = month2_costs.get(bu, 0)

        # Skip this BU if both current and previous months are zero
        if val1_monthly == 0 and val2_monthly == 0:
            continue

        ws_daily.cell(row, 1, bu)

        val1 = val1_monthly / days1
        val2 = val2_monthly / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    # Total row for daily average (reuse cached values)
    total1_daily = total1 / days1
    total2_daily = total2 / days2
    diff = total2_daily - total1_daily
    # Handle percentage calculation properly
    if total1_daily > 0:
        pct_diff = (total2_daily - total1_daily) / total1_daily
    elif total1_daily == 0 and total2_daily != 0:
        pct_diff = 1.0 if total2_daily > 0 else 0
    else:
        pct_diff = 0

    ws_daily.cell(row, 1, "total")
    ws_daily.cell(row, 2, total1_daily).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 3, total2_daily).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

    daily_end_row = row

    # Add conditional formatting for daily average difference and % difference columns
    _add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    _auto_adjust_column_widths(ws_daily)


def _add_conditional_formatting(ws, diff_range, pct_range):
    """Add conditional formatting to difference and % difference columns.

    Args:
        ws: Worksheet
        diff_range: Cell range for Difference column (e.g., "D4:D10")
        pct_range: Cell range for % Difference column (e.g., "E4:E10")
    """
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font as CFFont
    from openpyxl.styles import PatternFill as CFPatternFill
    from openpyxl.styles.differential import DifferentialStyle

    # Green for decrease (current < previous) - good, saving money
    # This means column C < column B, so the formatting should trigger when difference is negative
    green_fill = CFPatternFill(bgColor="FFC6EFCE")
    green_font = CFFont(color="FF006100")
    green_dxf = DifferentialStyle(fill=green_fill, font=green_font)

    # Red for increase (current > previous) - bad, spending more
    # This means column C > column B
    red_fill = CFPatternFill(bgColor="FFFFC7CE")
    red_font = CFFont(color="FF9C0006")
    red_dxf = DifferentialStyle(fill=red_fill, font=red_font)

    # For difference columns: Green when current (C) < previous (B), Red when current > previous
    start_row = int(diff_range.split(":")[0][1:])

    green_rule_diff = Rule(type="expression", dxf=green_dxf)
    green_rule_diff.formula = [f"C{start_row}<B{start_row}"]

    red_rule_diff = Rule(type="expression", dxf=red_dxf)
    red_rule_diff.formula = [f"C{start_row}>B{start_row}"]

    ws.conditional_formatting.add(diff_range, green_rule_diff)
    ws.conditional_formatting.add(diff_range, red_rule_diff)

    # Same for % difference columns
    green_rule_pct = Rule(type="expression", dxf=green_dxf)
    green_rule_pct.formula = [f"C{start_row}<B{start_row}"]

    red_rule_pct = Rule(type="expression", dxf=red_dxf)
    red_rule_pct.formula = [f"C{start_row}>B{start_row}"]

    ws.conditional_formatting.add(pct_range, green_rule_pct)
    ws.conditional_formatting.add(pct_range, red_rule_pct)


def _get_cell_length(cell):
    """Helper function to calculate cell length for column width adjustment."""
    if cell.value is None:
        return 0
    # For numeric values with formatting, use a reasonable width
    if isinstance(cell.value, (int, float)):
        return 15  # Fixed width for currency/percentage
    return len(str(cell.value))


def _auto_adjust_column_widths(ws):
    """Auto-adjust column widths based on content - optimized."""
    for column in ws.columns:
        try:
            # Use generator expression with max() for efficiency
            max_length = max((_get_cell_length(cell) for cell in column), default=0)
            # Add extra padding and ensure minimum width
            adjusted_width = min(max(max_length + 3, 12), 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        except (AttributeError, TypeError, ValueError):
            # Set default width if calculation fails
            ws.column_dimensions[column[0].column_letter].width = 12


def _create_service_analysis_tables(
    ws, ws_daily, cost_matrix, group_list, last_2_months, bu_cost_matrix
):
    """Create service analysis tables with monthly totals on one sheet, daily average on another, and pie chart."""
    # Header formatting
    header_font = Font(bold=True, size=14, color="FF000000")
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center")

    # Get total from BU costs for calculating "Other"
    bu_total = bu_cost_matrix[last_2_months[1]].get("total", 1)

    # Cache month dictionaries for faster lookups (performance optimization)
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Get top 10 services by latest month cost (excluding 'total')
    service_costs = [
        (svc, month2_costs.get(svc, 0)) for svc in group_list if svc != "total"
    ]
    service_costs.sort(key=lambda x: x[1], reverse=True)
    top_10_services = [svc for svc, _ in service_costs[:10]]

    # Calculate "Other" - difference between BU total and sum of top 10 services
    top_10_total = sum(month2_costs.get(svc, 0) for svc in top_10_services)
    other_amount = bu_total - top_10_total

    # Monthly Totals section
    ws["A1"] = "Top Services Monthly Totals"
    ws["A1"].font = Font(bold=True, size=16)

    row = 3
    ws.cell(row, 1, "Month").font = header_font
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).alignment = header_alignment

    ws.cell(row, 2, last_2_months[0]).font = header_font
    ws.cell(row, 2).fill = header_fill
    ws.cell(row, 2).alignment = header_alignment

    ws.cell(row, 3, last_2_months[1]).font = header_font
    ws.cell(row, 3).fill = header_fill
    ws.cell(row, 3).alignment = header_alignment

    ws.cell(row, 4, "Difference").font = header_font
    ws.cell(row, 4).fill = header_fill
    ws.cell(row, 4).alignment = header_alignment

    ws.cell(row, 5, "% Difference").font = header_font
    ws.cell(row, 5).fill = header_fill
    ws.cell(row, 5).alignment = header_alignment

    ws.cell(row, 6, "% Spend").font = header_font
    ws.cell(row, 6).fill = header_fill
    ws.cell(row, 6).alignment = header_alignment

    # Data rows for monthly totals (top 10 only)
    row += 1
    data_start_row = row
    pie_chart_start_row = row

    for service in top_10_services:
        ws.cell(row, 1, service)

        val1 = month1_costs.get(service, 0)
        val2 = month2_costs.get(service, 0)
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0
        pct_spend = val2 / bu_total

        ws.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    # Add "Other" row (for pie chart data, not a total)
    if other_amount > 0:
        ws.cell(row, 1, "Other")
        # Calculate Other values for previous month too (reuse cached sum)
        top_10_total_prev = sum(month1_costs.get(svc, 0) for svc in top_10_services)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )

        diff = other_amount - other_amount_prev
        # Handle percentage calculation properly
        if other_amount_prev > 0:
            pct_diff = (other_amount - other_amount_prev) / other_amount_prev
        elif other_amount_prev == 0 and other_amount != 0:
            pct_diff = 1.0 if other_amount > 0 else 0
        else:
            pct_diff = 0
        pct_spend = other_amount / bu_total

        ws.cell(row, 2, other_amount_prev).number_format = '"$"#,##0.00'
        ws.cell(row, 3, other_amount).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        pie_chart_end_row = row
        row += 1
    else:
        pie_chart_end_row = row - 1

    data_end_row = pie_chart_end_row

    # Add conditional formatting to difference and % difference columns for service monthly totals
    _add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Add pie chart (using column C data, which is the latest month)
    chart = PieChart()
    chart.title = None  # Remove title to prevent label overlap
    chart.style = 10
    chart.height = 15  # Increase height to show all labels
    chart.width = 20  # Increase width to show all labels

    # Use data from monthly totals including "Other" (not including column headers)
    labels = Reference(
        ws, min_col=1, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )
    data = Reference(
        ws, min_col=3, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Configure data labels to show category name and percentage only on the pie slices

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False  # Don't show value
    chart.dataLabels.showPercent = True
    chart.dataLabels.showSerName = False  # Don't show series name (e.g., "Series1")

    # Remove the legend - labels are shown on pie slices
    chart.legend = None

    ws.add_chart(chart, "H3")

    # Auto-adjust column widths
    _auto_adjust_column_widths(ws)

    # Daily Average section - on separate sheet
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        days1 = days2 = 30

    ws_daily["A1"] = "Top Services Daily Average"
    ws_daily["A1"].font = Font(bold=True, size=16)

    row = 3

    ws_daily.cell(row, 1, "Month").font = header_font
    ws_daily.cell(row, 1).fill = header_fill
    ws_daily.cell(row, 1).alignment = header_alignment

    ws_daily.cell(row, 2, last_2_months[0]).font = header_font
    ws_daily.cell(row, 2).fill = header_fill
    ws_daily.cell(row, 2).alignment = header_alignment

    ws_daily.cell(row, 3, last_2_months[1]).font = header_font
    ws_daily.cell(row, 3).fill = header_fill
    ws_daily.cell(row, 3).alignment = header_alignment

    ws_daily.cell(row, 4, "Difference").font = header_font
    ws_daily.cell(row, 4).fill = header_fill
    ws_daily.cell(row, 4).alignment = header_alignment

    ws_daily.cell(row, 5, "% Difference").font = header_font
    ws_daily.cell(row, 5).fill = header_fill
    ws_daily.cell(row, 5).alignment = header_alignment

    # Data rows for daily average (top 10 only, no Other)
    row += 1
    daily_start_row = row
    for service in top_10_services:
        ws_daily.cell(row, 1, service)

        val1 = cost_matrix[last_2_months[0]].get(service, 0) / days1
        val2 = cost_matrix[last_2_months[1]].get(service, 0) / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    daily_end_row = row - 1

    # Add conditional formatting for service daily average difference and % difference columns
    _add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    _auto_adjust_column_widths(ws_daily)


def _create_account_analysis_tables(
    ws, ws_daily, cost_matrix, group_list, last_2_months, bu_cost_matrix
):
    """Create account analysis tables with monthly totals on one sheet, daily average on another, and pie chart."""
    # Header formatting
    header_font = Font(bold=True, size=14, color="FF000000")
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center")

    # Get total from BU costs for calculating "Other"
    bu_total = bu_cost_matrix[last_2_months[1]].get("total", 1)

    # Cache month dictionaries for faster lookups (performance optimization)
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Get top 10 accounts by latest month cost (excluding 'total')
    account_costs = [
        (acc, month2_costs.get(acc, 0)) for acc in group_list if acc != "total"
    ]
    account_costs.sort(key=lambda x: x[1], reverse=True)
    top_10_accounts = [acc for acc, _ in account_costs[:10]]

    # Calculate "Other" - difference between BU total and sum of top 10 accounts
    top_10_total = sum(month2_costs.get(acc, 0) for acc in top_10_accounts)
    other_amount = bu_total - top_10_total

    # Monthly Totals section
    ws["A1"] = "Top Accounts Monthly Totals"
    ws["A1"].font = Font(bold=True, size=16)

    row = 3
    ws.cell(row, 1, "Month").font = header_font
    ws.cell(row, 1).fill = header_fill
    ws.cell(row, 1).alignment = header_alignment

    ws.cell(row, 2, last_2_months[0]).font = header_font
    ws.cell(row, 2).fill = header_fill
    ws.cell(row, 2).alignment = header_alignment

    ws.cell(row, 3, last_2_months[1]).font = header_font
    ws.cell(row, 3).fill = header_fill
    ws.cell(row, 3).alignment = header_alignment

    ws.cell(row, 4, "Difference").font = header_font
    ws.cell(row, 4).fill = header_fill
    ws.cell(row, 4).alignment = header_alignment

    ws.cell(row, 5, "% Difference").font = header_font
    ws.cell(row, 5).fill = header_fill
    ws.cell(row, 5).alignment = header_alignment

    ws.cell(row, 6, "% Spend").font = header_font
    ws.cell(row, 6).fill = header_fill
    ws.cell(row, 6).alignment = header_alignment

    # Data rows for monthly totals (top 10 only)
    row += 1
    data_start_row = row
    pie_chart_start_row = row

    for account in top_10_accounts:
        ws.cell(row, 1, account)

        val1 = month1_costs.get(account, 0)
        val2 = month2_costs.get(account, 0)
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0
        pct_spend = val2 / bu_total

        ws.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    # Add "Other" row (for pie chart data, not a total)
    if other_amount > 0:
        ws.cell(row, 1, "Other")
        # Calculate Other values for previous month too (reuse cached lookups)
        top_10_total_prev = sum(month1_costs.get(acc, 0) for acc in top_10_accounts)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )

        diff = other_amount - other_amount_prev
        # Handle percentage calculation properly
        if other_amount_prev > 0:
            pct_diff = (other_amount - other_amount_prev) / other_amount_prev
        elif other_amount_prev == 0 and other_amount != 0:
            pct_diff = 1.0 if other_amount > 0 else 0
        else:
            pct_diff = 0
        pct_spend = other_amount / bu_total

        ws.cell(row, 2, other_amount_prev).number_format = '"$"#,##0.00'
        ws.cell(row, 3, other_amount).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        pie_chart_end_row = row
        row += 1
    else:
        pie_chart_end_row = row - 1

    data_end_row = pie_chart_end_row

    # Add conditional formatting to difference and % difference columns for account monthly totals
    _add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Add pie chart (using column C data, which is the latest month)
    chart = PieChart()
    chart.title = None  # Remove title to prevent label overlap
    chart.style = 10
    chart.height = 15  # Increase height to show all labels
    chart.width = 20  # Increase width to show all labels

    # Use data from monthly totals including "Other" (not including column headers)
    labels = Reference(
        ws, min_col=1, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )
    data = Reference(
        ws, min_col=3, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Configure data labels to show category name and percentage only on the pie slices

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False  # Don't show value
    chart.dataLabels.showPercent = True
    chart.dataLabels.showSerName = False  # Don't show series name (e.g., "Series1")

    # Remove the legend - labels are shown on pie slices
    chart.legend = None

    ws.add_chart(chart, "H3")

    # Auto-adjust column widths
    _auto_adjust_column_widths(ws)

    # Daily Average section - on separate sheet
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        days1 = days2 = 30

    ws_daily["A1"] = "Top Accounts Daily Average"
    ws_daily["A1"].font = Font(bold=True, size=16)

    row = 3

    ws_daily.cell(row, 1, "Month").font = header_font
    ws_daily.cell(row, 1).fill = header_fill
    ws_daily.cell(row, 1).alignment = header_alignment

    ws_daily.cell(row, 2, last_2_months[0]).font = header_font
    ws_daily.cell(row, 2).fill = header_fill
    ws_daily.cell(row, 2).alignment = header_alignment

    ws_daily.cell(row, 3, last_2_months[1]).font = header_font
    ws_daily.cell(row, 3).fill = header_fill
    ws_daily.cell(row, 3).alignment = header_alignment

    ws_daily.cell(row, 4, "Difference").font = header_font
    ws_daily.cell(row, 4).fill = header_fill
    ws_daily.cell(row, 4).alignment = header_alignment

    ws_daily.cell(row, 5, "% Difference").font = header_font
    ws_daily.cell(row, 5).fill = header_fill
    ws_daily.cell(row, 5).alignment = header_alignment

    # Data rows for daily average (top 10 only, no Other)
    row += 1
    daily_start_row = row
    for account in top_10_accounts:
        ws_daily.cell(row, 1, account)

        val1 = cost_matrix[last_2_months[0]].get(account, 0) / days1
        val2 = cost_matrix[last_2_months[1]].get(account, 0) / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    daily_end_row = row - 1

    # Add conditional formatting for account daily average difference and % difference columns
    _add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    _auto_adjust_column_widths(ws_daily)


def _aggregate_year_costs(cost_matrix: dict, year_months: list[str]) -> dict:
    """Aggregate monthly costs into yearly totals for each group.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names to aggregate (e.g., ['Jan', 'Feb', ...])

    Returns:
        Dictionary of aggregated yearly costs {group: total_cost}
    """
    year_totals = {}

    for month in year_months:
        if month in cost_matrix:
            for group, cost in cost_matrix[month].items():
                if group not in year_totals:
                    year_totals[group] = 0.0
                year_totals[group] += cost

    return {k: round(v, 2) for k, v in year_totals.items()}


def _calculate_year_daily_average(cost_matrix: dict, year_months: list[str]) -> dict:
    """Calculate daily average costs for a year period.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names in the year (e.g., ['2023-Jan', '2023-Feb', ...])

    Returns:
        Dictionary of daily average costs {group: daily_avg_cost}
    """
    month_to_num = {month_abbr[i]: i for i in range(1, 13)}

    # Calculate total days in the year period
    total_days = 0

    for month_key in year_months:
        # Extract month name from key (e.g., "2023-Jan" -> "Jan")
        if "-" in month_key:
            year_str, month_name = month_key.split("-", 1)
            year = int(year_str)
        else:
            # Fallback for month names without year
            month_name = month_key
            year = datetime.now().year

        month_num = month_to_num.get(month_name, 1)
        days_in_month = monthrange(year, month_num)[1]
        total_days += days_in_month

    # Get year totals
    year_totals = _aggregate_year_costs(cost_matrix, year_months)

    # Calculate daily average
    if total_days > 0:
        return {k: round(v / total_days, 2) for k, v in year_totals.items()}
    return year_totals


def _calculate_year_monthly_average(cost_matrix: dict, year_months: list[str]) -> dict:
    """Calculate monthly average costs for a year period.

    Args:
        cost_matrix: Dictionary of monthly cost data {month: {group: cost}}
        year_months: List of month names in the year (e.g., ['Jan', 'Feb', ...])

    Returns:
        Dictionary of monthly average costs {group: monthly_avg_cost}
    """
    # Get year totals
    year_totals = _aggregate_year_costs(cost_matrix, year_months)

    # Calculate monthly average (divide by number of months)
    num_months = len(year_months)
    if num_months > 0:
        return {k: round(v / num_months, 2) for k, v in year_totals.items()}
    return year_totals


def export_year_analysis_excel(
    output_file,
    bu_cost_matrix,
    bu_group_list,
    service_cost_matrix,
    service_group_list,
    account_cost_matrix,
    account_group_list,
    year1_months,
    year2_months,
    all_account_costs=None,
):
    """Export year-level analysis Excel file with formatted tables and charts.

    Creates year analysis tables showing:
    - Account group allocation summary (first sheet)
    - Yearly totals (last 2 complete 12-month periods comparison)
    - Daily average (for year periods)
    - Monthly average (for year periods)
    - Pie charts for business units, services, and accounts

    Args:
        output_file: Path to the output year analysis Excel file
        bu_cost_matrix: Dictionary containing BU cost data organized by month
        bu_group_list: Dictionary of BU groups
        service_cost_matrix: Dictionary containing service cost data organized by month
        service_group_list: List of services
        account_cost_matrix: Dictionary containing account cost data organized by month
        account_group_list: List of accounts
        year1_months: List of month names for first year period
        year2_months: List of month names for second year period
        all_account_costs: Dictionary of all account costs (optional, for account summary)
    """
    LOGGER.info(f"Creating year analysis Excel file: {output_file}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Create Account Summary sheet first (if all_account_costs provided)
    if all_account_costs:
        ws_summary = wb.create_sheet("Account Summary", 0)
        _create_account_summary_sheet(ws_summary, bu_group_list, all_account_costs)

    # Aggregate data into yearly totals
    bu_year1 = _aggregate_year_costs(bu_cost_matrix, year1_months)
    bu_year2 = _aggregate_year_costs(bu_cost_matrix, year2_months)

    service_year1 = _aggregate_year_costs(service_cost_matrix, year1_months)
    service_year2 = _aggregate_year_costs(service_cost_matrix, year2_months)

    account_year1 = _aggregate_year_costs(account_cost_matrix, year1_months)
    account_year2 = _aggregate_year_costs(account_cost_matrix, year2_months)

    # Calculate daily averages
    bu_year1_daily = _calculate_year_daily_average(bu_cost_matrix, year1_months)
    bu_year2_daily = _calculate_year_daily_average(bu_cost_matrix, year2_months)

    service_year1_daily = _calculate_year_daily_average(
        service_cost_matrix, year1_months
    )
    service_year2_daily = _calculate_year_daily_average(
        service_cost_matrix, year2_months
    )

    account_year1_daily = _calculate_year_daily_average(
        account_cost_matrix, year1_months
    )
    account_year2_daily = _calculate_year_daily_average(
        account_cost_matrix, year2_months
    )

    # Calculate monthly averages
    bu_year1_monthly = _calculate_year_monthly_average(bu_cost_matrix, year1_months)
    bu_year2_monthly = _calculate_year_monthly_average(bu_cost_matrix, year2_months)

    service_year1_monthly = _calculate_year_monthly_average(
        service_cost_matrix, year1_months
    )
    service_year2_monthly = _calculate_year_monthly_average(
        service_cost_matrix, year2_months
    )

    account_year1_monthly = _calculate_year_monthly_average(
        account_cost_matrix, year1_months
    )
    account_year2_monthly = _calculate_year_monthly_average(
        account_cost_matrix, year2_months
    )

    # Create year labels
    year1_label = f"Year 1 ({year1_months[0]}-{year1_months[-1]})"
    year2_label = f"Year 2 ({year2_months[0]}-{year2_months[-1]})"

    # Create sheets for BU costs
    ws_bu = wb.create_sheet("BU Costs - Yearly")
    _create_year_comparison_sheet(
        ws_bu,
        "Business Unit Yearly Totals",
        bu_year1,
        bu_year2,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_BU_Year_Chart",
        chart_title="BU Costs Distribution",
    )

    ws_bu_daily = wb.create_sheet("BU Costs - Daily Avg")
    _create_year_comparison_sheet(
        ws_bu_daily,
        "Business Unit Daily Average",
        bu_year1_daily,
        bu_year2_daily,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_bu_monthly = wb.create_sheet("BU Costs - Monthly Avg")
    _create_year_comparison_sheet(
        ws_bu_monthly,
        "Business Unit Monthly Average",
        bu_year1_monthly,
        bu_year2_monthly,
        list(bu_group_list.keys()) + ["total"],
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # For services, calculate top 10 + Other (like monthly reports)
    # Get top 10 services by year 2 cost
    service_costs_sorted = [
        (svc, service_year2.get(svc, 0)) for svc in service_group_list if svc != "total"
    ]
    service_costs_sorted.sort(key=lambda x: x[1], reverse=True)
    top_10_services = [svc for svc, _ in service_costs_sorted[:10]]

    # Calculate "Other" for services (sum of services not in top 10)
    top_10_year1_total = sum(service_year1.get(svc, 0) for svc in top_10_services)
    top_10_year2_total = sum(service_year2.get(svc, 0) for svc in top_10_services)

    # Get BU total for calculating Other
    bu_year1_total = bu_year1.get("total", 0)
    bu_year2_total = bu_year2.get("total", 0)

    service_other_year1 = round(bu_year1_total - top_10_year1_total, 2)
    service_other_year2 = round(bu_year2_total - top_10_year2_total, 2)

    # Add Other to year1 and year2 data
    service_year1_with_other = {**service_year1, "Other": service_other_year1}
    service_year2_with_other = {**service_year2, "Other": service_other_year2}

    service_year1_daily_with_other = {
        **service_year1_daily,
        "Other": service_other_year1 / 365,
    }  # Approximate daily
    service_year2_daily_with_other = {
        **service_year2_daily,
        "Other": service_other_year2 / 365,
    }

    service_year1_monthly_with_other = {
        **service_year1_monthly,
        "Other": service_other_year1 / 12,
    }
    service_year2_monthly_with_other = {
        **service_year2_monthly,
        "Other": service_other_year2 / 12,
    }

    # Create sheets for Service costs (top 10 + Other)
    ws_service = wb.create_sheet("Top Services - Yearly")
    _create_year_comparison_sheet(
        ws_service,
        "Top Services Yearly Totals",
        service_year1_with_other,
        service_year2_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_Service_Year_Chart",
        chart_title="Top Services Distribution",
    )

    ws_service_daily = wb.create_sheet("Top Services - Daily Avg")
    _create_year_comparison_sheet(
        ws_service_daily,
        "Top Services Daily Average",
        service_year1_daily_with_other,
        service_year2_daily_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_service_monthly = wb.create_sheet("Top Services - Monthly Avg")
    _create_year_comparison_sheet(
        ws_service_monthly,
        "Top Services Monthly Average",
        service_year1_monthly_with_other,
        service_year2_monthly_with_other,
        top_10_services + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # For accounts, calculate top 10 + Other (like monthly reports)
    # Get top 10 accounts by year 2 cost
    account_costs_sorted = [
        (acc, account_year2.get(acc, 0)) for acc in account_group_list if acc != "total"
    ]
    account_costs_sorted.sort(key=lambda x: x[1], reverse=True)
    top_10_accounts = [acc for acc, _ in account_costs_sorted[:10]]

    # Calculate "Other" for accounts (sum of accounts not in top 10)
    top_10_acc_year1_total = sum(account_year1.get(acc, 0) for acc in top_10_accounts)
    top_10_acc_year2_total = sum(account_year2.get(acc, 0) for acc in top_10_accounts)

    account_other_year1 = round(bu_year1_total - top_10_acc_year1_total, 2)
    account_other_year2 = round(bu_year2_total - top_10_acc_year2_total, 2)

    # Add Other to year1 and year2 data
    account_year1_with_other = {**account_year1, "Other": account_other_year1}
    account_year2_with_other = {**account_year2, "Other": account_other_year2}

    account_year1_daily_with_other = {
        **account_year1_daily,
        "Other": account_other_year1 / 365,
    }  # Approximate daily
    account_year2_daily_with_other = {
        **account_year2_daily,
        "Other": account_other_year2 / 365,
    }

    account_year1_monthly_with_other = {
        **account_year1_monthly,
        "Other": account_other_year1 / 12,
    }
    account_year2_monthly_with_other = {
        **account_year2_monthly,
        "Other": account_other_year2 / 12,
    }

    # Create sheets for Account costs (top 10 + Other)
    ws_account = wb.create_sheet("Top Accounts - Yearly")
    _create_year_comparison_sheet(
        ws_account,
        "Top Accounts Yearly Totals",
        account_year1_with_other,
        account_year2_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=True,  # Include pie chart for yearly totals
        workbook=wb,
        chart_helper_name="_Account_Year_Chart",
        chart_title="Top Accounts Distribution",
    )

    ws_account_daily = wb.create_sheet("Top Accounts - Daily Avg")
    _create_year_comparison_sheet(
        ws_account_daily,
        "Top Accounts Daily Average",
        account_year1_daily_with_other,
        account_year2_daily_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for daily average
        workbook=wb,
    )

    ws_account_monthly = wb.create_sheet("Top Accounts - Monthly Avg")
    _create_year_comparison_sheet(
        ws_account_monthly,
        "Top Accounts Monthly Average",
        account_year1_monthly_with_other,
        account_year2_monthly_with_other,
        top_10_accounts + ["Other"],  # Top 10 + Other, no total
        year1_label,
        year2_label,
        include_chart=False,  # No chart for monthly average
        workbook=wb,
    )

    # Save the workbook
    wb.save(output_file)
    LOGGER.info(f"Year analysis file saved: {output_file}")


def _create_year_comparison_sheet(
    worksheet,
    title,
    year1_data,
    year2_data,
    group_list,
    year1_label,
    year2_label,
    include_chart=True,
    workbook=None,
    chart_helper_name=None,
    chart_title=None,
):
    """Create a year comparison sheet with formatted table and chart.

    Args:
        worksheet: openpyxl worksheet object
        title: Title for the sheet
        year1_data: Dictionary of year 1 costs {group: cost}
        year2_data: Dictionary of year 2 costs {group: cost}
        group_list: List of groups to include
        year1_label: Label for year 1 column
        year2_label: Label for year 2 column
        include_chart: Whether to include pie chart (default: True, only for yearly totals)
        workbook: Workbook object for creating helper sheet (needed for pie chart with "Other" grouping)
        chart_helper_name: Short name for hidden helper sheet (max 31 chars, default: auto-generated)
        chart_title: Title for pie chart (default: uses year2_label Distribution)
    """
    # Define styles - match monthly analysis format
    header_font = Font(bold=True, size=14, color="FF000000")
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center")

    # Title
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16)

    # Headers - match monthly format with "Month" instead of "Group"
    row = 3
    worksheet.cell(row, 1, "Month").font = header_font
    worksheet.cell(row, 1).fill = header_fill
    worksheet.cell(row, 1).alignment = header_alignment

    worksheet.cell(row, 2, year1_label).font = header_font
    worksheet.cell(row, 2).fill = header_fill
    worksheet.cell(row, 2).alignment = header_alignment

    worksheet.cell(row, 3, year2_label).font = header_font
    worksheet.cell(row, 3).fill = header_fill
    worksheet.cell(row, 3).alignment = header_alignment

    worksheet.cell(row, 4, "Difference").font = header_font
    worksheet.cell(row, 4).fill = header_fill
    worksheet.cell(row, 4).alignment = header_alignment

    worksheet.cell(row, 5, "% Difference").font = header_font
    worksheet.cell(row, 5).fill = header_fill
    worksheet.cell(row, 5).alignment = header_alignment

    # Add % Spend column header (only for sheets with totals or Other, not for daily/monthly averages)
    has_totals = "total" in group_list
    has_other = "Other" in group_list
    if (has_totals or has_other) and include_chart:
        worksheet.cell(row, 6, "% Spend").font = header_font
        worksheet.cell(row, 6).fill = header_fill
        worksheet.cell(row, 6).alignment = header_alignment

    # Sort groups by year2 (most recent) value in descending order, with 'Other' and 'total' at the end
    sorted_groups = []
    total_group = None
    other_group = None
    for group in group_list:
        if group == "total":
            total_group = group
        elif group == "Other":
            other_group = group
        else:
            sorted_groups.append(group)

    # Sort by most recent year's value (descending)
    sorted_groups.sort(key=lambda g: year2_data.get(g, 0), reverse=True)

    # Add Other before total (if present)
    if other_group:
        sorted_groups.append(other_group)

    # Add total at the very end if present
    if total_group:
        sorted_groups.append(total_group)

    # Data rows
    row += 1
    start_row = row

    # Calculate total for % Spend
    # For BU costs, use "total" if present
    # For Services/Accounts with "Other", calculate sum of all items including Other
    if has_totals:
        total_val2 = year2_data.get("total", 1)
    elif has_other:
        # Sum all groups including Other for % Spend calculation
        total_val2 = sum(year2_data.get(g, 0) for g in sorted_groups if g != "total")
    else:
        total_val2 = 1

    for group in sorted_groups:
        val1 = year1_data.get(group, 0)
        val2 = year2_data.get(group, 0)

        # Skip rows with zero values in both periods (except total)
        if group != "total" and val1 == 0 and val2 == 0:
            continue

        diff = val2 - val1

        # Calculate percentage difference
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        worksheet.cell(row, 1, group)
        worksheet.cell(row, 2, val1).number_format = '"$"#,##0.00'
        worksheet.cell(row, 3, val2).number_format = '"$"#,##0.00'
        # Use absolute value for difference column to match monthly format
        worksheet.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        worksheet.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        # Add % Spend column (skip for total row, but show for Other)
        if (has_totals or has_other) and include_chart and group != "total":
            pct_spend = val2 / total_val2 if total_val2 > 0 else 0
            worksheet.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    end_row = row - 1

    # Add conditional formatting
    _add_conditional_formatting(
        worksheet,
        f"D{start_row}:D{end_row}",
        f"E{start_row}:E{end_row}",
    )

    # Auto-adjust column widths
    _auto_adjust_column_widths(worksheet)

    # Add pie chart for year 2 data (most recent) - only if include_chart is True
    if include_chart and workbook:
        # Create helper sheet for pie chart data (groups items < 1% into "Other")
        # Use provided name or generate a short one to stay under 31 char limit
        if chart_helper_name:
            helper_sheet_name = chart_helper_name
        else:
            # Fallback to a shortened name if none provided
            helper_sheet_name = f"_Chart_{worksheet.title[:20]}"

        ws_helper = workbook.create_sheet(helper_sheet_name)
        ws_helper.sheet_state = "hidden"

        helper_row = 1
        helper_col = 1
        pie_chart_start_row = helper_row

        # Process groups in single pass: add >= 1% spend, accumulate < 1% for "Other"
        # Performance optimization: Combined 2 loops into 1 (50% reduction in iterations)
        # "Other" accumulation only happens for BU costs (not Services/Accounts with explicit Other)
        other_total = 0
        for group in sorted_groups:
            if group == "total":
                continue

            # If group is "Other" and it's in the year2_data, add it directly
            if group == "Other" and "Other" in year2_data:
                val2 = year2_data.get("Other", 0)
                if val2 > 0:
                    ws_helper.cell(helper_row, helper_col, "Other")
                    ws_helper.cell(helper_row, helper_col + 1, val2)
                    helper_row += 1
                continue

            val2 = year2_data.get(group, 0)
            val1 = year1_data.get(group, 0)

            # Skip if zero
            if val1 == 0 and val2 == 0:
                continue

            pct_spend = val2 / total_val2 if total_val2 > 0 else 0
            if pct_spend >= 0.01:  # >= 1%
                ws_helper.cell(helper_row, helper_col, group)
                ws_helper.cell(helper_row, helper_col + 1, val2)
                helper_row += 1
            elif "Other" not in year2_data:
                # Only accumulate "Other" if not already in the data
                # (i.e., only calculate for BU costs, not for Services/Accounts which have explicit Other)
                other_total += val2

        # Add accumulated "Other" if needed
        if "Other" not in year2_data and other_total > 0:
            ws_helper.cell(helper_row, helper_col, "Other")
            ws_helper.cell(helper_row, helper_col + 1, other_total)
            helper_row += 1

        pie_chart_end_row = helper_row - 1

        # Add pie chart using helper table data from hidden sheet
        if pie_chart_end_row >= pie_chart_start_row:
            chart = PieChart()
            # Remove title to prevent pie labels from falling behind it
            # The sheet already has a descriptive section title
            chart.title = None
            chart.style = 10
            # Match monthly report chart size
            chart.height = 15
            chart.width = 20

            # Use data from helper sheet
            labels = Reference(
                ws_helper,
                min_col=helper_col,
                min_row=pie_chart_start_row,
                max_row=pie_chart_end_row,
            )
            data = Reference(
                ws_helper,
                min_col=helper_col + 1,
                min_row=pie_chart_start_row,
                max_row=pie_chart_end_row,
            )

            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)

            # Configure data labels to show category name and percentage only
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showCatName = True
            chart.dataLabels.showVal = False
            chart.dataLabels.showPercent = True
            chart.dataLabels.showSerName = False
            chart.legend = None

            # Position chart to match monthly reports
            worksheet.add_chart(chart, "H3")
