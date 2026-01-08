"""Report export functions for CSV and Excel formats.

This module provides functions for exporting cost reports to CSV and Excel formats.
"""

import csv

from openpyxl import Workbook

from amc.reportexport.formatting import EXPORT_HEADER_FILL, EXPORT_HEADER_FONT


def export_report(
    export_file, cost_matrix, group_list, group_by_type, output_format="csv"
):
    """Export cost report to CSV or Excel format.

    Args:
        export_file: Path to the output file
        cost_matrix: Dictionary containing cost data organized by month
        group_list: List or dictionary of groups (accounts, BUs, or services)
        group_by_type: Type of grouping ("account", "bu", or "service")
        output_format: Output format, either "csv" or "excel" (default: "csv")
    """
    # Directory should be created by caller, but ensure it exists
    export_file.parent.mkdir(parents=True, exist_ok=True)

    months = list(cost_matrix.keys())

    if output_format == "excel":
        _export_to_excel(export_file, cost_matrix, group_list, group_by_type, months)
    else:
        _export_to_csv(export_file, cost_matrix, group_list, group_by_type, months)


def _export_to_csv(export_file, cost_matrix, group_list, group_by_type, months):
    """Export cost report to CSV format."""
    csv_header = ["Month"] + months

    with open(export_file, "w", newline="") as ef:
        writer = csv.writer(ef)
        writer.writerow(csv_header)

        if group_by_type == "account":
            for account in group_list:
                csv_row = [account] + [
                    cost_matrix[month].get(account, 0) for month in months
                ]
                writer.writerow(csv_row)
        elif group_by_type == "bu":
            # Get all BUs from config and any additional ones in data (like "unallocated")
            first_month = list(cost_matrix.keys())[0] if cost_matrix else None
            if first_month:
                bus_from_config = set(group_list.keys())
                bus_from_data = set(cost_matrix[first_month].keys()) - {"total"}
                bus = list(bus_from_config | bus_from_data) + ["total"]
            else:
                bus = list(group_list.keys()) + ["total"]
            for bu in bus:
                csv_row = [bu] + [cost_matrix[month].get(bu, 0) for month in months]
                writer.writerow(csv_row)
        elif group_by_type == "service":
            for service in group_list:
                csv_row = [service] + [
                    cost_matrix[month].get(service, 0) for month in months
                ]
                writer.writerow(csv_row)


def _export_to_excel(export_file, cost_matrix, group_list, group_by_type, months):
    """Export cost report to Excel format."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "AWS Costs"

    # Write header
    header = ["Month"] + months
    for col_idx, header_value in enumerate(header, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header_value)
        cell.font = EXPORT_HEADER_FONT
        cell.fill = EXPORT_HEADER_FILL

    # Write data rows
    row_idx = 2
    if group_by_type == "account":
        for account in group_list:
            worksheet.cell(row=row_idx, column=1, value=account)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(account, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
    elif group_by_type == "bu":
        # Get all BUs from config and any additional ones in data (like "unallocated")
        first_month = months[0] if months else None
        if first_month:
            bus_from_config = set(group_list.keys())
            bus_from_data = set(cost_matrix[first_month].keys()) - {"total"}
            bus = list(bus_from_config | bus_from_data) + ["total"]
        else:
            bus = list(group_list.keys()) + ["total"]
        for bu in bus:
            worksheet.cell(row=row_idx, column=1, value=bu)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(bu, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
    elif group_by_type == "service":
        for service in group_list:
            worksheet.cell(row=row_idx, column=1, value=service)
            for col_idx, month in enumerate(months, start=2):
                value = cost_matrix[month].get(service, 0)
                worksheet.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

    # Auto-adjust column widths - optimized with single pass
    for col_idx, column in enumerate(worksheet.columns, start=1):
        try:
            # Use generator expression with max() for efficiency
            max_length = max(
                (len(str(cell.value)) for cell in column if cell.value is not None),
                default=0,
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except (AttributeError, TypeError, ValueError):
            # Set default width if calculation fails
            worksheet.column_dimensions[column[0].column_letter].width = 12

    # Save workbook
    workbook.save(export_file)
