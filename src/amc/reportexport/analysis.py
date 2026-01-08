"""Analysis table creation functions for Excel reports.

This module provides functions for creating detailed analysis tables with
monthly/daily comparisons, pie charts, and conditional formatting.
"""

import logging
from calendar import monthrange
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from amc.reportexport.calculations import (
    calculate_difference,
    calculate_percentage_difference,
    calculate_percentage_spend,
)
from amc.reportexport.formatting import (
    add_conditional_formatting,
    apply_currency_format,
    apply_header_style,
    apply_percentage_format,
    auto_adjust_column_widths,
    auto_adjust_column_widths_advanced,
)
from amc.reportexport.charts import (
    add_chart_to_worksheet,
    add_data_to_pie_chart,
    create_pie_chart,
)
from amc.reportexport.analysis_tables import (
    create_analysis_header_row,
    create_section_title,
    get_top_n_items,
)

LOGGER = logging.getLogger(__name__)


def _create_account_summary_sheet(
    worksheet,
    account_groups,
    all_account_costs,
    account_id_to_name=None,
    comparison_months=None,
):
    """Create account group allocation summary sheet.

    Args:
        worksheet: openpyxl worksheet object
        account_groups: Dictionary of business unit account groups
        all_account_costs: Dictionary of all account costs from API
        account_id_to_name: Optional dictionary mapping account IDs to names
        comparison_months: Optional list of months to filter accounts by (show only accounts with costs in these months)
    """
    # Get all account IDs that are defined in account_groups
    assigned_accounts = set()
    for bu, bu_accounts in account_groups.items():
        assigned_accounts.update(bu_accounts.keys())

    # Get account IDs from cost data, filtered by comparison months if provided
    if comparison_months:
        # Only include accounts that have costs in at least one of the comparison months
        all_cost_account_ids = set()
        for month in comparison_months:
            if month in all_account_costs:
                all_cost_account_ids.update(all_account_costs[month].keys())
    else:
        # Use first month as before for backward compatibility
        first_month_key = next(iter(all_account_costs.keys()))
        all_cost_account_ids = set(all_account_costs[first_month_key].keys())

    # Identify unallocated accounts
    unallocated_account_ids = all_cost_account_ids - assigned_accounts

    # Define styles matching existing sheets
    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14, color="FF000000")
    header_fill = PatternFill(
        start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center")
    warning_font = Font(bold=True, size=12, color="FF9C0006")
    warning_fill = PatternFill(
        start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"
    )

    # Title
    worksheet["A1"] = "ACCOUNT GROUP ALLOCATION SUMMARY"
    worksheet["A1"].font = title_font

    row = 3

    # Show accounts by BU
    for bu in sorted(account_groups.keys()):
        bu_accounts = account_groups[bu]

        # BU header with styling
        cell = worksheet.cell(row, 1, f"{bu.upper()}")
        cell.font = section_font
        row += 1

        # Column headers with styling matching existing sheets
        header_cell_1 = worksheet.cell(row, 1, "Account ID")
        header_cell_1.font = header_font
        header_cell_1.fill = header_fill
        header_cell_1.alignment = header_alignment

        if account_id_to_name:
            header_cell_2 = worksheet.cell(row, 2, "Account Name")
            header_cell_2.font = header_font
            header_cell_2.fill = header_fill
            header_cell_2.alignment = header_alignment
        row += 1

        if bu_accounts:
            # Filter accounts to only show those with costs in comparison periods (if specified)
            accounts_to_show = []
            for account_id in sorted(bu_accounts.keys()):
                # If filtering by comparison months, only show accounts with costs in those months
                if comparison_months:
                    has_costs = any(
                        month in all_account_costs
                        and account_id in all_account_costs[month]
                        for month in comparison_months
                    )
                    if has_costs:
                        accounts_to_show.append(account_id)
                else:
                    accounts_to_show.append(account_id)

            if accounts_to_show:
                for account_id in accounts_to_show:
                    worksheet.cell(row, 1, account_id)
                    if account_id_to_name and account_id in account_id_to_name:
                        worksheet.cell(row, 2, account_id_to_name[account_id])
                    row += 1
            else:
                cell = worksheet.cell(
                    row, 1, "(no accounts with costs in comparison period)"
                )
                cell.font = Font(italic=True, color="FF999999")
                row += 1
        else:
            cell = worksheet.cell(row, 1, "(no accounts)")
            cell.font = Font(italic=True, color="FF999999")
            row += 1

        row += 1  # Add blank line between BUs

    # Show unallocated accounts if any with warning styling
    unalloc_header = worksheet.cell(
        row, 1, f"UNALLOCATED ACCOUNTS ({len(unallocated_account_ids)})"
    )
    if unallocated_account_ids:
        unalloc_header.font = warning_font
        unalloc_header.fill = warning_fill
    else:
        unalloc_header.font = section_font
    row += 1

    if unallocated_account_ids:
        # Column headers with styling
        header_cell_1 = worksheet.cell(row, 1, "Account ID")
        header_cell_1.font = header_font
        header_cell_1.fill = header_fill
        header_cell_1.alignment = header_alignment

        if account_id_to_name:
            header_cell_2 = worksheet.cell(row, 2, "Account Name")
            header_cell_2.font = header_font
            header_cell_2.fill = header_fill
            header_cell_2.alignment = header_alignment
        row += 1

        for account_id in sorted(unallocated_account_ids):
            worksheet.cell(row, 1, account_id)
            if account_id_to_name and account_id in account_id_to_name:
                worksheet.cell(row, 2, account_id_to_name[account_id])
            row += 1
    else:
        cell = worksheet.cell(row, 1, "None")
        cell.font = Font(italic=True, color="FF666666")
        row += 1

    # Auto-adjust column widths
    worksheet.column_dimensions["A"].width = 20
    if account_id_to_name:
        worksheet.column_dimensions["B"].width = 40


def export_analysis_excel(
    output_file,
    bu_cost_matrix,
    bu_group_list,
    service_cost_matrix,
    service_group_list,
    account_cost_matrix,
    account_group_list,
    all_account_costs=None,
    account_id_to_name=None,
):
    """Export analysis Excel file with formatted tables and pie charts.

    Creates analysis tables showing:
    - Account group allocation summary (first sheet)
    - Monthly totals (last 2 months comparison)
    - Daily average (last 2 months comparison)
    - Pie charts for business units, services, and accounts

    Args:
        output_file: Path to the output analysis Excel file
        bu_cost_matrix: Dictionary containing BU cost data organized by month
        bu_group_list: Dictionary of BU groups
        service_cost_matrix: Dictionary containing service cost data organized by month
        service_group_list: List of services
        account_cost_matrix: Dictionary containing account cost data organized by month
        account_group_list: List of accounts
        all_account_costs: Dictionary of all account costs (optional, for account summary)
        account_id_to_name: Dictionary mapping account IDs to names (optional, for account summary)
    """

    LOGGER.info(f"Creating analysis Excel file: {output_file}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Get last 2 months from the data - sort chronologically
    def get_sort_key(month_str):
        """Generate a sortable key for month strings.

        Handles both simple month format (e.g., 'Jan') and YYYY-Mon format (e.g., '2023-Jan').
        For simple format, assumes current/recent year for sorting.
        """
        try:
            # Try simple month format first (e.g., "Jan")
            month_date = datetime.strptime(month_str, "%b")
            # For simple format, use a recent year for sorting
            return datetime(2024, month_date.month, 1)
        except ValueError:
            try:
                # Try YYYY-Mon format (e.g., "2023-Jan")
                return datetime.strptime(month_str, "%Y-%b")
            except ValueError:
                # Fallback to string sorting
                return datetime.min

    months = sorted(list(bu_cost_matrix.keys()), key=get_sort_key)
    if len(months) < 2:
        LOGGER.warning("Need at least 2 months of data for analysis")
        return

    last_2_months = months[-2:]

    # Create Account Summary sheet first (if all_account_costs provided)
    if all_account_costs:
        ws_summary = wb.create_sheet("Account Summary", 0)
        _create_account_summary_sheet(
            ws_summary,
            bu_group_list,
            all_account_costs,
            account_id_to_name,
            last_2_months,
        )

    # Create analysis sheets
    ws_bu = wb.create_sheet("BU Costs")
    ws_bu_daily = wb.create_sheet("BU Daily Average")
    ws_bu_helper = wb.create_sheet("_BU_Chart_Data")
    ws_bu_helper.sheet_state = "hidden"  # Hide the helper sheet
    _create_bu_analysis_tables(
        ws_bu, ws_bu_daily, ws_bu_helper, bu_cost_matrix, bu_group_list, last_2_months
    )

    ws_service = wb.create_sheet("Top Services")
    ws_service_daily = wb.create_sheet("Top Services Daily Avg")
    _create_service_analysis_tables(
        ws_service,
        ws_service_daily,
        service_cost_matrix,
        service_group_list,
        last_2_months,
        bu_cost_matrix,
    )

    ws_account = wb.create_sheet("Top Accounts")
    ws_account_daily = wb.create_sheet("Top Accounts Daily Avg")
    _create_account_analysis_tables(
        ws_account,
        ws_account_daily,
        account_cost_matrix,
        account_group_list,
        last_2_months,
        bu_cost_matrix,
    )

    # Ensure output directory exists
    output_file.parent.mkdir(parents=True, exist_ok=True)

    # Save the workbook
    wb.save(output_file)
    LOGGER.info(f"Analysis Excel file saved: {output_file}")


def _create_bu_analysis_tables(
    ws, ws_daily, ws_helper, cost_matrix, group_list, last_2_months
):
    """Create BU analysis tables with monthly totals on one sheet, daily average on another, and pie chart with helper table in hidden sheet."""
    # Monthly Totals section
    ws["A1"] = "BU Monthly Totals"
    ws["A1"].font = Font(bold=True, size=16)

    row = 3
    # Apply header styling using utility functions
    for col, header_text in enumerate(
        [
            "Month",
            last_2_months[0],
            last_2_months[1],
            "Difference",
            "% Difference",
            "% Spend",
        ],
        start=1,
    ):
        apply_header_style(ws.cell(row, col, header_text))

    # Data rows for monthly totals
    row += 1
    data_start_row = row

    # Cache month dictionaries for faster lookups (performance optimization)
    # These cached references eliminate repeated dictionary traversals in the loops below
    # for main table, daily average calculations, and chart data processing
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]
    month2_total = month2_costs.get("total", 1)

    # Get all BUs from both group_list and cost_matrix (to include "unallocated")
    # Exclude 'total' as it's handled separately
    bus_from_config = set(group_list.keys())
    bus_from_data = set(month2_costs.keys()) - {"total"}
    all_bus = list(bus_from_config | bus_from_data)

    # Sort BUs by most recent month's cost in descending order
    all_bus.sort(key=lambda bu: month2_costs.get(bu, 0), reverse=True)

    # Display all BUs in the table
    for bu in all_bus:
        val1 = month1_costs.get(bu, 0)
        val2 = month2_costs.get(bu, 0)

        # Skip this BU if both current and previous months are zero
        if val1 == 0 and val2 == 0:
            continue

        ws.cell(row, 1, bu)

        # Use calculation utilities
        diff = calculate_difference(val1, val2)
        pct_diff = calculate_percentage_difference(val1, val2)
        pct_spend = calculate_percentage_spend(val2, month2_total)

        # Apply formatting using utility functions
        apply_currency_format(ws.cell(row, 2, val1))
        apply_currency_format(ws.cell(row, 3, val2))
        apply_currency_format(ws.cell(row, 4, diff))
        apply_percentage_format(ws.cell(row, 5, abs(pct_diff)))
        apply_percentage_format(ws.cell(row, 6, pct_spend))

        row += 1

    # Total row
    total1 = month1_costs.get("total", 0)
    total2 = month2_costs.get("total", 0)
    diff = calculate_difference(total1, total2)
    pct_diff = calculate_percentage_difference(total1, total2)

    ws.cell(row, 1, "total")
    apply_currency_format(ws.cell(row, 2, total1))
    apply_currency_format(ws.cell(row, 3, total2))
    apply_currency_format(ws.cell(row, 4, diff))
    apply_percentage_format(ws.cell(row, 5, abs(pct_diff)))
    # Column 6 (% Spend) intentionally left empty for total row - it's implied to be 100%

    data_end_row = row

    # Add conditional formatting to difference and % difference columns
    add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Create helper table for pie chart in hidden sheet (groups BUs < 1% into "Other")
    helper_row = 1
    helper_col = 1  # Column A in helper sheet

    total_for_pct = month2_total  # Reuse cached value
    pie_chart_start_row = helper_row

    # Process BUs in single pass: add >= 1% spend, accumulate < 1% for "Other"
    # Performance optimization: Combined 2 loops into 1 (50% reduction in iterations)
    other_total = 0
    for bu in all_bus:
        val2 = month2_costs.get(bu, 0)
        val1 = month1_costs.get(bu, 0)

        # Skip if zero
        if val1 == 0 and val2 == 0:
            continue

        pct_spend = val2 / total_for_pct
        if pct_spend >= 0.01:  # >= 1%
            ws_helper.cell(helper_row, helper_col, bu)
            ws_helper.cell(helper_row, helper_col + 1, val2)
            helper_row += 1
        else:  # < 1%
            other_total += val2

    if other_total > 0:
        ws_helper.cell(helper_row, helper_col, "Other")
        ws_helper.cell(helper_row, helper_col + 1, other_total)
        helper_row += 1

    pie_chart_end_row = helper_row - 1

    # Add pie chart using helper table data from hidden sheet
    # Only add pie chart if there's BU data to display
    if pie_chart_end_row >= pie_chart_start_row:
        chart = create_pie_chart(show_legend=False, show_series_name=False)
        chart = add_data_to_pie_chart(
            chart,
            ws_helper,
            helper_col + 1,
            helper_col,
            pie_chart_start_row,
            pie_chart_end_row,
        )
        add_chart_to_worksheet(ws, chart, "H3")

    # Auto-adjust column widths
    auto_adjust_column_widths(ws)

    # Daily Average section - on separate sheet
    ws_daily["A1"] = "BU Daily Average"
    ws_daily["A1"].font = Font(bold=True, size=16)

    row = 3

    # Apply header styling using utility functions
    for col, header_text in enumerate(
        ["Month", last_2_months[0], last_2_months[1], "Difference", "% Difference"],
        start=1,
    ):
        apply_header_style(ws_daily.cell(row, col, header_text))

    # Calculate days in each month
    # For the last 2 months, infer the year intelligently
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        # Fallback to 30 days if parsing fails
        days1 = days2 = 30

    # Data rows for daily average
    row += 1
    daily_start_row = row
    for bu in all_bus:
        val1_monthly = month1_costs.get(bu, 0)
        val2_monthly = month2_costs.get(bu, 0)

        # Skip this BU if both current and previous months are zero
        if val1_monthly == 0 and val2_monthly == 0:
            continue

        ws_daily.cell(row, 1, bu)

        val1 = val1_monthly / days1
        val2 = val2_monthly / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    # Total row for daily average (reuse cached values)
    total1_daily = total1 / days1
    total2_daily = total2 / days2
    diff = total2_daily - total1_daily
    # Handle percentage calculation properly
    if total1_daily > 0:
        pct_diff = (total2_daily - total1_daily) / total1_daily
    elif total1_daily == 0 and total2_daily != 0:
        pct_diff = 1.0 if total2_daily > 0 else 0
    else:
        pct_diff = 0

    ws_daily.cell(row, 1, "total")
    ws_daily.cell(row, 2, total1_daily).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 3, total2_daily).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
    ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

    daily_end_row = row

    # Add conditional formatting for daily average difference and % difference columns
    add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(ws_daily)


def _create_service_analysis_tables(
    ws, ws_daily, cost_matrix, group_list, last_2_months, bu_cost_matrix
):
    """Create service analysis tables with monthly totals on one sheet, daily average on another, and pie chart."""
    # Get total from BU costs for calculating "Other"
    bu_total = bu_cost_matrix[last_2_months[1]].get("total", 1)

    # Cache month dictionaries for faster lookups (performance optimization)
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Get top 10 services by latest month cost (excluding 'total')
    top_10_services = get_top_n_items(cost_matrix, group_list, last_2_months, n=10)

    # Calculate "Other" - difference between BU total and sum of top 10 services
    top_10_total = sum(month2_costs.get(svc, 0) for svc in top_10_services)
    other_amount = bu_total - top_10_total

    # Monthly Totals section
    create_section_title(ws, "Top Services Monthly Totals")

    row = 3
    headers = [
        "Month",
        last_2_months[0],
        last_2_months[1],
        "Difference",
        "% Difference",
        "% Spend",
    ]
    create_analysis_header_row(ws, row, headers)

    # Data rows for monthly totals (top 10 only)
    row += 1
    data_start_row = row
    pie_chart_start_row = row

    for service in top_10_services:
        ws.cell(row, 1, service)

        val1 = month1_costs.get(service, 0)
        val2 = month2_costs.get(service, 0)
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0
        pct_spend = val2 / bu_total

        ws.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    # Add "Other" row (for pie chart data, not a total)
    if other_amount > 0:
        ws.cell(row, 1, "Other")
        # Calculate Other values for previous month too (reuse cached sum)
        top_10_total_prev = sum(month1_costs.get(svc, 0) for svc in top_10_services)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )

        diff = other_amount - other_amount_prev
        # Handle percentage calculation properly
        if other_amount_prev > 0:
            pct_diff = (other_amount - other_amount_prev) / other_amount_prev
        elif other_amount_prev == 0 and other_amount != 0:
            pct_diff = 1.0 if other_amount > 0 else 0
        else:
            pct_diff = 0
        pct_spend = other_amount / bu_total

        ws.cell(row, 2, other_amount_prev).number_format = '"$"#,##0.00'
        ws.cell(row, 3, other_amount).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        pie_chart_end_row = row
        row += 1
    else:
        pie_chart_end_row = row - 1

    data_end_row = pie_chart_end_row

    # Add conditional formatting to difference and % difference columns for service monthly totals
    add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Add pie chart (using column C data, which is the latest month)
    chart = PieChart()
    chart.title = None  # Remove title to prevent label overlap
    chart.style = 10
    chart.height = 15  # Increase height to show all labels
    chart.width = 20  # Increase width to show all labels

    # Use data from monthly totals including "Other" (not including column headers)
    labels = Reference(
        ws, min_col=1, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )
    data = Reference(
        ws, min_col=3, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Configure data labels to show category name and percentage only on the pie slices

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False  # Don't show value
    chart.dataLabels.showPercent = True
    chart.dataLabels.showSerName = False  # Don't show series name (e.g., "Series1")

    # Remove the legend - labels are shown on pie slices
    chart.legend = None

    ws.add_chart(chart, "H3")

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(ws)

    # Daily Average section - on separate sheet
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        days1 = days2 = 30

    create_section_title(ws_daily, "Top Services Daily Average")

    row = 3
    headers = [
        "Month",
        last_2_months[0],
        last_2_months[1],
        "Difference",
        "% Difference",
    ]
    create_analysis_header_row(ws_daily, row, headers)

    # Data rows for daily average (top 10 only)
    row += 1
    daily_start_row = row
    for service in top_10_services:
        ws_daily.cell(row, 1, service)

        val1 = cost_matrix[last_2_months[0]].get(service, 0) / days1
        val2 = cost_matrix[last_2_months[1]].get(service, 0) / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    # Add "Other" row for daily average
    if other_amount > 0:
        ws_daily.cell(row, 1, "Other")
        # Calculate Other daily average values for both months
        top_10_total_prev = sum(month1_costs.get(svc, 0) for svc in top_10_services)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )
        val1 = other_amount_prev / days1
        val2 = other_amount / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    daily_end_row = row - 1

    # Add conditional formatting for service daily average difference and % difference columns
    add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(ws_daily)


def _create_account_analysis_tables(
    ws, ws_daily, cost_matrix, group_list, last_2_months, bu_cost_matrix
):
    """Create account analysis tables with monthly totals on one sheet, daily average on another, and pie chart."""
    # Get total from BU costs for calculating "Other"
    bu_total = bu_cost_matrix[last_2_months[1]].get("total", 1)

    # Cache month dictionaries for faster lookups (performance optimization)
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Get top 10 accounts by latest month cost (excluding 'total')
    top_10_accounts = get_top_n_items(cost_matrix, group_list, last_2_months, n=10)

    # Calculate "Other" - difference between BU total and sum of top 10 accounts
    top_10_total = sum(month2_costs.get(acc, 0) for acc in top_10_accounts)
    other_amount = bu_total - top_10_total

    # Monthly Totals section
    create_section_title(ws, "Top Accounts Monthly Totals")

    row = 3
    headers = [
        "Month",
        last_2_months[0],
        last_2_months[1],
        "Difference",
        "% Difference",
        "% Spend",
    ]
    create_analysis_header_row(ws, row, headers)

    # Data rows for monthly totals (top 10 only)
    row += 1
    data_start_row = row
    pie_chart_start_row = row

    for account in top_10_accounts:
        ws.cell(row, 1, account)

        val1 = month1_costs.get(account, 0)
        val2 = month2_costs.get(account, 0)
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0
        pct_spend = val2 / bu_total

        ws.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        row += 1

    # Add "Other" row (for pie chart data, not a total)
    if other_amount > 0:
        ws.cell(row, 1, "Other")
        # Calculate Other values for previous month too (reuse cached lookups)
        top_10_total_prev = sum(month1_costs.get(acc, 0) for acc in top_10_accounts)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )

        diff = other_amount - other_amount_prev
        # Handle percentage calculation properly
        if other_amount_prev > 0:
            pct_diff = (other_amount - other_amount_prev) / other_amount_prev
        elif other_amount_prev == 0 and other_amount != 0:
            pct_diff = 1.0 if other_amount > 0 else 0
        else:
            pct_diff = 0
        pct_spend = other_amount / bu_total

        ws.cell(row, 2, other_amount_prev).number_format = '"$"#,##0.00'
        ws.cell(row, 3, other_amount).number_format = '"$"#,##0.00'
        ws.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws.cell(row, 5, abs(pct_diff)).number_format = "0.00%"
        ws.cell(row, 6, pct_spend).number_format = "0.00%"

        pie_chart_end_row = row
        row += 1
    else:
        pie_chart_end_row = row - 1

    data_end_row = pie_chart_end_row

    # Add conditional formatting to difference and % difference columns for account monthly totals
    add_conditional_formatting(
        ws, f"D{data_start_row}:D{data_end_row}", f"E{data_start_row}:E{data_end_row}"
    )

    # Add pie chart (using column C data, which is the latest month)
    chart = PieChart()
    chart.title = None  # Remove title to prevent label overlap
    chart.style = 10
    chart.height = 15  # Increase height to show all labels
    chart.width = 20  # Increase width to show all labels

    # Use data from monthly totals including "Other" (not including column headers)
    labels = Reference(
        ws, min_col=1, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )
    data = Reference(
        ws, min_col=3, min_row=pie_chart_start_row, max_row=pie_chart_end_row
    )

    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)

    # Configure data labels to show category name and percentage only on the pie slices

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False  # Don't show value
    chart.dataLabels.showPercent = True
    chart.dataLabels.showSerName = False  # Don't show series name (e.g., "Series1")

    # Remove the legend - labels are shown on pie slices
    chart.legend = None

    ws.add_chart(chart, "H3")

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(ws)

    # Daily Average section - on separate sheet
    try:
        month1_date = datetime.strptime(last_2_months[0], "%b")
        month2_date = datetime.strptime(last_2_months[1], "%b")

        # Infer years for the last 2 months
        # If month2 < month1 (e.g., Dec -> Jan), they span year boundary
        current_year = datetime.now().year
        if month2_date.month < month1_date.month:
            # Year boundary case: month1 is from previous year
            year1 = current_year - 1
            year2 = current_year
        else:
            # Normal case: both months from same year
            # Determine if they're from current or previous year
            current_month = datetime.now().month
            if month2_date.month <= current_month:
                # Both months are from current year or earlier
                year1 = current_year
                year2 = current_year
            else:
                # Both months are from previous year
                year1 = current_year - 1
                year2 = current_year - 1

        days1 = monthrange(year1, month1_date.month)[1]
        days2 = monthrange(year2, month2_date.month)[1]
    except ValueError:
        days1 = days2 = 30

    create_section_title(ws_daily, "Top Accounts Daily Average")

    row = 3
    headers = [
        "Month",
        last_2_months[0],
        last_2_months[1],
        "Difference",
        "% Difference",
    ]
    create_analysis_header_row(ws_daily, row, headers)

    # Data rows for daily average (top 10 only)
    row += 1
    daily_start_row = row
    for account in top_10_accounts:
        ws_daily.cell(row, 1, account)

        val1 = cost_matrix[last_2_months[0]].get(account, 0) / days1
        val2 = cost_matrix[last_2_months[1]].get(account, 0) / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    # Add "Other" row for daily average
    if other_amount > 0:
        ws_daily.cell(row, 1, "Other")
        # Calculate Other daily average values for both months
        top_10_total_prev = sum(month1_costs.get(acc, 0) for acc in top_10_accounts)
        other_amount_prev = (
            bu_cost_matrix[last_2_months[0]].get("total", 0) - top_10_total_prev
        )
        val1 = other_amount_prev / days1
        val2 = other_amount / days2
        diff = val2 - val1
        # Handle percentage calculation properly
        if val1 > 0:
            pct_diff = (val2 - val1) / val1
        elif val1 == 0 and val2 != 0:
            pct_diff = 1.0 if val2 > 0 else 0
        else:
            pct_diff = 0

        ws_daily.cell(row, 2, val1).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 3, val2).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 4, abs(diff)).number_format = '"$"#,##0.00'
        ws_daily.cell(row, 5, abs(pct_diff)).number_format = "0.00%"

        row += 1

    daily_end_row = row - 1

    # Add conditional formatting for account daily average difference and % difference columns
    add_conditional_formatting(
        ws_daily,
        f"D{daily_start_row}:D{daily_end_row}",
        f"E{daily_start_row}:E{daily_end_row}",
    )

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(ws_daily)
