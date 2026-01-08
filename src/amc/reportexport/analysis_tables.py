"""Analysis table utilities for Excel reports.

This module provides reusable functions for creating analysis tables with
monthly totals, daily averages, and pie charts.
"""

from openpyxl.styles import Font

from amc.reportexport.calculations import calculate_percentage_difference
from amc.reportexport.formatting import (
    CURRENCY_FORMAT,
    HEADER_ALIGNMENT_CENTER,
    HEADER_FILL_STANDARD,
    HEADER_FONT_STANDARD,
    PERCENTAGE_FORMAT,
    add_conditional_formatting,
    auto_adjust_column_widths_advanced,
)
from amc.reportexport.charts import (
    add_chart_to_worksheet,
    add_data_to_pie_chart,
    create_pie_chart,
)


def get_top_n_items(cost_matrix, group_list, last_2_months, n=10):
    """Get top N items by latest month cost.

    Args:
        cost_matrix: Dictionary containing cost data organized by month
        group_list: List of items (services, accounts, etc.)
        last_2_months: List of the last 2 months
        n: Number of top items to return (default: 10)

    Returns:
        List of top N item names
    """
    month2_costs = cost_matrix[last_2_months[1]]
    item_costs = [
        (item, month2_costs.get(item, 0)) for item in group_list if item != "total"
    ]
    item_costs.sort(key=lambda x: x[1], reverse=True)
    return [item for item, _ in item_costs[:n]]


def calculate_other_amount(cost_matrix, top_items, last_2_months, bu_cost_matrix):
    """Calculate the 'Other' amount (total minus top N).

    Args:
        cost_matrix: Dictionary containing cost data organized by month
        top_items: List of top N items
        last_2_months: List of the last 2 months
        bu_cost_matrix: Business unit cost matrix containing totals

    Returns:
        Tuple of (other_amount_latest, other_amount_previous)
    """
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Get BU totals from the BU cost matrix
    bu_total_latest = bu_cost_matrix[last_2_months[1]].get("total", 0)
    bu_total_prev = bu_cost_matrix[last_2_months[0]].get("total", 0)

    top_total_latest = sum(month2_costs.get(item, 0) for item in top_items)
    other_amount_latest = bu_total_latest - top_total_latest

    top_total_prev = sum(month1_costs.get(item, 0) for item in top_items)
    other_amount_prev = bu_total_prev - top_total_prev

    return other_amount_latest, other_amount_prev


def create_analysis_header_row(worksheet, row, headers):
    """Create a header row with standard formatting.

    Args:
        worksheet: The worksheet to write to
        row: Row number for the header
        headers: List of header text values

    Returns:
        None (modifies worksheet in place)
    """
    for col, header_text in enumerate(headers, start=1):
        cell = worksheet.cell(row, col, header_text)
        cell.font = HEADER_FONT_STANDARD
        cell.fill = HEADER_FILL_STANDARD
        cell.alignment = HEADER_ALIGNMENT_CENTER


def create_section_title(worksheet, title, cell="A1"):
    """Create a section title with standard formatting.

    Args:
        worksheet: The worksheet to write to
        title: Title text
        cell: Cell location for the title (default: "A1")

    Returns:
        None (modifies worksheet in place)
    """
    worksheet[cell] = title
    worksheet[cell].font = Font(bold=True, size=16)


def write_data_row(worksheet, row, item_name, val1, val2, bu_total):
    """Write a single data row with costs, difference, and percentages.

    Args:
        worksheet: The worksheet to write to
        row: Row number for the data
        item_name: Name of the item (service, account, etc.)
        val1: Cost for previous month
        val2: Cost for latest month
        bu_total: Total BU cost for percentage calculation

    Returns:
        None (modifies worksheet in place)
    """
    worksheet.cell(row, 1, item_name)

    diff = val2 - val1
    pct_diff = calculate_percentage_difference(val1, val2)
    pct_spend = val2 / bu_total if bu_total > 0 else 0

    worksheet.cell(row, 2, val1).number_format = CURRENCY_FORMAT
    worksheet.cell(row, 3, val2).number_format = CURRENCY_FORMAT
    worksheet.cell(row, 4, abs(diff)).number_format = CURRENCY_FORMAT
    worksheet.cell(row, 5, abs(pct_diff)).number_format = PERCENTAGE_FORMAT
    worksheet.cell(row, 6, pct_spend).number_format = PERCENTAGE_FORMAT


def create_monthly_totals_table(
    worksheet, cost_matrix, items, last_2_months, bu_cost_matrix, title, include_other=True
):
    """Create a monthly totals table with top items.

    Args:
        worksheet: The worksheet to write to
        cost_matrix: Dictionary containing cost data organized by month
        items: List of items to include in the table
        last_2_months: List of the last 2 months
        bu_cost_matrix: Business unit cost matrix containing totals
        title: Title for the table
        include_other: Whether to include an "Other" row (default: True)

    Returns:
        Tuple of (end_row, pie_chart_start_row) for further processing
    """
    # Cache month dictionaries for faster lookups
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]
    bu_total = bu_cost_matrix[last_2_months[1]].get("total", 1)

    # Title
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16)

    # Header row
    row = 3
    headers = ["Month", last_2_months[0], last_2_months[1], "Difference", "% Difference", "% Spend"]
    create_analysis_header_row(worksheet, row, headers)

    # Data rows
    row += 1
    data_start_row = row
    pie_chart_start_row = row

    for item in items:
        val1 = month1_costs.get(item, 0)
        val2 = month2_costs.get(item, 0)
        write_data_row(worksheet, row, item, val1, val2, bu_total)
        row += 1

    # Add "Other" row if requested
    if include_other:
        other_latest, other_prev = calculate_other_amount(
            cost_matrix, items, last_2_months, bu_cost_matrix
        )
        if other_latest > 0:
            write_data_row(worksheet, row, "Other", other_prev, other_latest, bu_total)
            row += 1

    # Add conditional formatting
    diff_range = f"D{data_start_row}:D{row - 1}"
    pct_range = f"E{data_start_row}:E{row - 1}"
    add_conditional_formatting(worksheet, diff_range, pct_range)

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(worksheet)

    return row, pie_chart_start_row


def create_daily_average_table(
    worksheet, cost_matrix, items, last_2_months, bu_total_daily, title
):
    """Create a daily average table with top items.

    Args:
        worksheet: The worksheet to write to
        cost_matrix: Dictionary containing daily average cost data
        items: List of items to include in the table
        last_2_months: List of the last 2 months
        bu_total_daily: Total BU daily average for percentage calculation
        title: Title for the table

    Returns:
        None (modifies worksheet in place)
    """
    # Cache month dictionaries for faster lookups
    month1_costs = cost_matrix[last_2_months[0]]
    month2_costs = cost_matrix[last_2_months[1]]

    # Title
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16)

    # Header row
    row = 3
    headers = ["Month", last_2_months[0], last_2_months[1], "Difference", "% Difference"]
    create_analysis_header_row(worksheet, row, headers)

    # Data rows
    row += 1
    data_start_row = row

    for item in items:
        val1 = month1_costs.get(item, 0)
        val2 = month2_costs.get(item, 0)

        worksheet.cell(row, 1, item)
        diff = val2 - val1
        pct_diff = calculate_percentage_difference(val1, val2)

        worksheet.cell(row, 2, val1).number_format = CURRENCY_FORMAT
        worksheet.cell(row, 3, val2).number_format = CURRENCY_FORMAT
        worksheet.cell(row, 4, abs(diff)).number_format = CURRENCY_FORMAT
        worksheet.cell(row, 5, abs(pct_diff)).number_format = PERCENTAGE_FORMAT

        row += 1

    # Add conditional formatting
    diff_range = f"D{data_start_row}:D{row - 1}"
    pct_range = f"E{data_start_row}:E{row - 1}"
    add_conditional_formatting(worksheet, diff_range, pct_range)

    # Auto-adjust column widths
    auto_adjust_column_widths_advanced(worksheet)


def create_pie_chart_with_data(
    worksheet_chart, worksheet_data, items, cost_matrix, last_month, title, chart_position
):
    """Create a pie chart with helper data for top items.

    Args:
        worksheet_chart: The worksheet to place the chart on
        worksheet_data: The hidden worksheet for chart data
        items: List of items to include
        cost_matrix: Dictionary containing cost data
        last_month: The month to use for the chart
        title: Title for the chart
        chart_position: Cell position for the chart (e.g., "H3")

    Returns:
        None (modifies worksheets in place)
    """
    month_costs = cost_matrix[last_month]

    # Prepare chart helper data with single-pass loop
    chart_data = []
    other_total = 0

    for item in items:
        cost = month_costs.get(item, 0)
        pct_spend = cost / month_costs.get("total", 1) if month_costs.get("total", 0) > 0 else 0

        if pct_spend >= 0.01:  # Include items >= 1%
            chart_data.append((item, cost))
        else:
            other_total += cost

    # Add "Other" if it exists
    if other_total > 0:
        chart_data.append(("Other", other_total))

    # Write data to hidden helper sheet
    row = 1
    for item_name, cost in chart_data:
        worksheet_data.cell(row, 1, item_name)
        worksheet_data.cell(row, 2, cost)
        row += 1

    # Create and configure pie chart
    chart = create_pie_chart(title=title)
    data_range = (1, 1, len(chart_data), 2)
    add_data_to_pie_chart(chart, worksheet_data, data_range)
    add_chart_to_worksheet(worksheet_chart, chart, chart_position)
