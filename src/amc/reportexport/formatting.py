"""Formatting utilities for Excel reports.

This module provides common formatting functions and styles used across
Excel report generation to eliminate code duplication.
"""

from openpyxl.styles import Alignment, Font, PatternFill


# Standard header styles
HEADER_FONT_STANDARD = Font(bold=True, size=14, color="FF000000")
HEADER_FILL_STANDARD = PatternFill(
    start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
)
HEADER_ALIGNMENT_CENTER = Alignment(horizontal="center")

# Year analysis header styles
HEADER_FONT_YEAR = Font(bold=True, size=14, color="FFFFFFFF")
HEADER_FILL_YEAR = PatternFill(
    start_color="FF4472C4", end_color="FF4472C4", fill_type="solid"
)

# Excel/CSV export header styles
EXPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
EXPORT_HEADER_FILL = PatternFill(
    start_color="366092", end_color="366092", fill_type="solid"
)

# Number format constants
CURRENCY_FORMAT = '"$"#,##0.00'
PERCENTAGE_FORMAT = "0.00%"


def create_header_font(bold=True, size=14, color="FF000000"):
    """Create a Font object for headers.

    Args:
        bold: Whether the font should be bold (default: True)
        size: Font size (default: 14)
        color: Font color in hex format (default: black)

    Returns:
        Font object
    """
    return Font(bold=bold, size=size, color=color)


def create_header_fill(color="FFD9E1F2"):
    """Create a PatternFill object for headers.

    Args:
        color: Fill color in hex format (default: light blue)

    Returns:
        PatternFill object
    """
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def apply_header_style(cell, font=None, fill=None, alignment=None):
    """Apply header styling to a cell.

    Args:
        cell: The cell to style
        font: Font object (default: standard header font)
        fill: PatternFill object (default: standard header fill)
        alignment: Alignment object (default: center alignment)

    Returns:
        The styled cell
    """
    cell.font = font or HEADER_FONT_STANDARD
    cell.fill = fill or HEADER_FILL_STANDARD
    cell.alignment = alignment or HEADER_ALIGNMENT_CENTER
    return cell


def apply_currency_format(cell):
    """Apply currency format to a cell.

    Args:
        cell: The cell to format

    Returns:
        The formatted cell
    """
    cell.number_format = CURRENCY_FORMAT
    return cell


def apply_percentage_format(cell):
    """Apply percentage format to a cell.

    Args:
        cell: The cell to format

    Returns:
        The formatted cell
    """
    cell.number_format = PERCENTAGE_FORMAT
    return cell


def auto_adjust_column_widths(worksheet, max_width=50, min_width=12):
    """Auto-adjust column widths based on content.

    Modifies the worksheet in place by adjusting column widths based on cell content.

    Args:
        worksheet: The worksheet to adjust
        max_width: Maximum column width (default: 50)
        min_width: Minimum column width (default: 12)

    Returns:
        None (modifies worksheet in place)
    """
    for col_idx, column in enumerate(worksheet.columns, start=1):
        try:
            # Use generator expression with max() for efficiency
            max_length = max(
                (len(str(cell.value)) for cell in column if cell.value is not None),
                default=0,
            )
            adjusted_width = min(max_length + 2, max_width)
            adjusted_width = max(adjusted_width, min_width)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except (AttributeError, TypeError, ValueError):
            # Set default width if calculation fails
            worksheet.column_dimensions[column[0].column_letter].width = min_width


def get_cell_length(cell):
    """Calculate cell length for column width adjustment.

    Args:
        cell: The cell to measure

    Returns:
        The calculated length for width adjustment
    """
    if cell.value is None:
        return 0
    # For numeric values with formatting, use a reasonable width
    if isinstance(cell.value, (int, float)):
        return 15  # Fixed width for currency/percentage
    return len(str(cell.value))


def auto_adjust_column_widths_advanced(worksheet):
    """Auto-adjust column widths based on content with advanced sizing.

    Uses get_cell_length for more accurate width calculations.

    Args:
        worksheet: The worksheet to adjust

    Returns:
        None (modifies worksheet in place)
    """
    for column in worksheet.columns:
        try:
            # Use generator expression with max() for efficiency
            max_length = max((get_cell_length(cell) for cell in column), default=0)
            # Add extra padding and ensure minimum width
            adjusted_width = min(max(max_length + 3, 12), 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except (AttributeError, TypeError, ValueError):
            # Set default width if calculation fails
            worksheet.column_dimensions[column[0].column_letter].width = 12


def add_conditional_formatting(worksheet, diff_range, pct_range):
    """Add conditional formatting to difference and percentage difference columns.

    Green is applied when current month (column C) is less than previous month (column B),
    indicating cost savings. Red is applied when current month is greater than previous month.

    Args:
        worksheet: The worksheet to apply formatting to
        diff_range: Cell range for Difference column (e.g., "D4:D10")
        pct_range: Cell range for % Difference column (e.g., "E4:E10")

    Returns:
        None (modifies worksheet in place)
    """
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles import Font as CFFont
    from openpyxl.styles import PatternFill as CFPatternFill
    from openpyxl.styles.differential import DifferentialStyle

    # Green for decrease (current < previous) - good, saving money
    green_fill = CFPatternFill(bgColor="FFC6EFCE")
    green_font = CFFont(color="FF006100")
    green_dxf = DifferentialStyle(fill=green_fill, font=green_font)

    # Red for increase (current > previous) - bad, spending more
    red_fill = CFPatternFill(bgColor="FFFFC7CE")
    red_font = CFFont(color="FF9C0006")
    red_dxf = DifferentialStyle(fill=red_fill, font=red_font)

    # For difference columns: Green when current (C) < previous (B), Red when current > previous
    start_row = int(diff_range.split(":")[0][1:])

    green_rule_diff = Rule(type="expression", dxf=green_dxf)
    green_rule_diff.formula = [f"C{start_row}<B{start_row}"]

    red_rule_diff = Rule(type="expression", dxf=red_dxf)
    red_rule_diff.formula = [f"C{start_row}>B{start_row}"]

    worksheet.conditional_formatting.add(diff_range, green_rule_diff)
    worksheet.conditional_formatting.add(diff_range, red_rule_diff)

    # Same for % difference columns
    green_rule_pct = Rule(type="expression", dxf=green_dxf)
    green_rule_pct.formula = [f"C{start_row}<B{start_row}"]

    red_rule_pct = Rule(type="expression", dxf=red_dxf)
    red_rule_pct.formula = [f"C{start_row}>B{start_row}"]

    worksheet.conditional_formatting.add(pct_range, green_rule_pct)
    worksheet.conditional_formatting.add(pct_range, red_rule_pct)
